from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Request, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import RedirectResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timedelta
import jwt
import bcrypt
from bson import ObjectId
import resend
import stripe
import httpx
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT_DIR = Path(__file__).parent

# Configure logging first before anything else with more verbose output
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    force=True  # Force reconfiguration if already configured
)
logger = logging.getLogger(__name__)

# Log startup
logger.info("=" * 60)
logger.info("AÏLA APPLICATION STARTING")
logger.info("=" * 60)

# Try to load .env file if it exists (for local development)
env_file = ROOT_DIR / '.env'
if env_file.exists():
    load_dotenv(env_file)
    logger.info(f"✓ Loaded environment variables from {env_file}")
else:
    logger.info("✓ No .env file found, using environment variables from system")
    logger.info(f"✓ Environment variables: MONGO_URL={'SET' if os.environ.get('MONGO_URL') else 'NOT SET'}, DB_NAME={os.environ.get('DB_NAME', 'NOT SET')}")

# Google OAuth Configuration
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '548263066328-916g23gmboqvmqtd7fi3ejatoseh4h09.apps.googleusercontent.com')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
if GOOGLE_CLIENT_SECRET:
    logger.info(f"✓ GOOGLE_CLIENT_SECRET configured ({len(GOOGLE_CLIENT_SECRET)} chars)")
else:
    logger.warning("⚠ GOOGLE_CLIENT_SECRET not set - Google OAuth will fail!")

# Configure Stripe
STRIPE_SECRET_KEY = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_PUBLISHABLE_KEY = os.environ.get('STRIPE_PUBLISHABLE_KEY', '')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
stripe.api_key = STRIPE_SECRET_KEY

# Stripe Price IDs (will be created dynamically if not set)
STRIPE_PRICE_MONTHLY = os.environ.get('STRIPE_PRICE_MONTHLY', '')
STRIPE_PRICE_YEARLY = os.environ.get('STRIPE_PRICE_YEARLY', '')
STRIPE_PRICE_LIFETIME = os.environ.get('STRIPE_PRICE_LIFETIME', '')

# Configure Resend for email sending
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', 're_ZVuwLNzR_ER7hfFTBQi6LqZTZuWEQa7Sr')
resend.api_key = RESEND_API_KEY
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'https://www.aila.family')

async def send_invitation_email(to_email: str, inviter_name: str, role: str, invite_token: str):
    """Send invitation email via Resend"""
    try:
        role_fr = "lecteur" if role == "viewer" else "éditeur"
        invite_url = f"{FRONTEND_URL}/(tabs)/tree?invite={invite_token}"
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #0A1628; color: #FFFFFF;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #D4AF37; margin: 0;">🌳 AÏLA</h1>
                <p style="color: #B8C5D6; margin-top: 5px;">L'arbre qui connecte les générations</p>
            </div>
            
            <div style="background-color: #1A2F4A; border-radius: 12px; padding: 24px; margin-bottom: 20px;">
                <h2 style="color: #FFFFFF; margin-top: 0;">Vous êtes invité(e) !</h2>
                <p style="color: #B8C5D6; line-height: 1.6;">
                    <strong style="color: #D4AF37;">{inviter_name}</strong> vous invite à rejoindre son arbre généalogique familial en tant que <strong style="color: #D4AF37;">{role_fr}</strong>.
                </p>
            </div>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{invite_url}" style="display: inline-block; background-color: #D4AF37; color: #0A1628; text-decoration: none; padding: 16px 32px; border-radius: 12px; font-weight: bold; font-size: 16px;">
                    Accepter l'invitation
                </a>
            </div>
            
            <p style="color: #6B7C93; font-size: 12px; text-align: center;">
                Si vous ne connaissez pas {inviter_name}, vous pouvez ignorer cet email.
            </p>
        </div>
        """
        
        params = {
            "from": "AÏLA <noreply@aila.family>",
            "to": [to_email],
            "subject": f"🌳 {inviter_name} vous invite à rejoindre son arbre généalogique",
            "html": html_content
        }
        
        email = resend.Emails.send(params)
        logger.info(f"✓ Invitation email sent to {to_email}")
        return True
    except Exception as e:
        logger.error(f"✗ Failed to send invitation email to {to_email}: {e}")
        return False

async def send_password_reset_email(to_email: str, reset_token: str):
    """Send password reset email via Resend"""
    try:
        reset_url = f"{FRONTEND_URL}/reset-password?token={reset_token}"
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #0A1628; color: #FFFFFF;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #D4AF37; margin: 0;">🌳 AÏLA</h1>
                <p style="color: #B8C5D6; margin-top: 5px;">Réinitialisation de mot de passe</p>
            </div>
            
            <div style="background-color: #1A2F4A; border-radius: 12px; padding: 24px; margin-bottom: 20px;">
                <h2 style="color: #FFFFFF; margin-top: 0;">Mot de passe oublié ?</h2>
                <p style="color: #B8C5D6; line-height: 1.6;">
                    Vous avez demandé la réinitialisation de votre mot de passe. Cliquez sur le bouton ci-dessous pour en créer un nouveau.
                </p>
                <p style="color: #B8C5D6; line-height: 1.6;">
                    <strong style="color: #D4AF37;">Ce lien expire dans 1 heure.</strong>
                </p>
            </div>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_url}" style="display: inline-block; background-color: #D4AF37; color: #0A1628; text-decoration: none; padding: 16px 32px; border-radius: 12px; font-weight: bold; font-size: 16px;">
                    Réinitialiser mon mot de passe
                </a>
            </div>
            
            <p style="color: #6B7C93; font-size: 12px; text-align: center;">
                Si vous n'avez pas demandé cette réinitialisation, ignorez cet email. Votre mot de passe restera inchangé.
            </p>
        </div>
        """
        
        params = {
            "from": "AÏLA <noreply@aila.family>",
            "to": [to_email],
            "subject": "🔑 Réinitialisation de votre mot de passe AÏLA",
            "html": html_content
        }
        
        email = resend.Emails.send(params)
        logger.info(f"✓ Password reset email sent to {to_email}")
        return True
    except Exception as e:
        logger.error(f"✗ Failed to send password reset email to {to_email}: {e}")
        return False

async def send_event_notification_email(to_email: str, sender_name: str, event_type: str, event_title: str, event_description: str = None, event_date: str = None):
    """Send event notification email via Resend"""
    try:
        # Emoji based on event type
        event_emojis = {
            "birthday": "🎂",
            "birth": "👶",
            "graduation": "🎓",
            "wedding": "💍",
            "newyear": "🎆",
            "holiday": "🎄",
            "custom": "🎉"
        }
        emoji = event_emojis.get(event_type, "🎉")
        
        # Event type label in French
        event_labels = {
            "birthday": "Anniversaire",
            "birth": "Naissance",
            "graduation": "Diplôme",
            "wedding": "Mariage",
            "newyear": "Nouvel An",
            "holiday": "Fête",
            "custom": "Événement"
        }
        event_label = event_labels.get(event_type, "Événement")
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #0A1628; color: #FFFFFF;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #D4AF37; margin: 0;">🌳 AÏLA</h1>
                <p style="color: #B8C5D6; margin-top: 5px;">Événement familial</p>
            </div>
            
            <div style="background-color: #1A2F4A; border-radius: 12px; padding: 24px; margin-bottom: 20px; text-align: center;">
                <div style="font-size: 48px; margin-bottom: 16px;">{emoji}</div>
                <h2 style="color: #D4AF37; margin: 0 0 16px 0;">{event_title}</h2>
                <p style="color: #B8C5D6; margin: 0;">
                    <strong>{sender_name}</strong> vous partage cet événement
                </p>
            </div>
            
            {f'<div style="background-color: #2A3F5A; border-radius: 8px; padding: 16px; margin-bottom: 20px;"><p style="color: #FFFFFF; margin: 0;">{event_description}</p></div>' if event_description else ''}
            
            {f'<p style="color: #6B7C93; text-align: center;">📅 {event_date}</p>' if event_date else ''}
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{FRONTEND_URL}" style="display: inline-block; background-color: #D4AF37; color: #0A1628; text-decoration: none; padding: 16px 32px; border-radius: 12px; font-weight: bold; font-size: 16px;">
                    Voir l'arbre familial
                </a>
            </div>
        </div>
        """
        
        params = {
            "from": "AÏLA <noreply@aila.family>",
            "to": [to_email],
            "subject": f"{emoji} {event_label} - {event_title}",
            "html": html_content
        }
        
        email = resend.Emails.send(params)
        logger.info(f"✓ Event notification email sent to {to_email}")
        return True
    except Exception as e:
        logger.error(f"✗ Failed to send event notification email to {to_email}: {e}")
        return False

# MongoDB connection with timeout settings for Atlas
logger.info("Configuring MongoDB connection...")
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    logger.warning("⚠ MONGO_URL environment variable not set, using default localhost")
    mongo_url = "mongodb://localhost:27017"
else:
    # Log a sanitized version (hide credentials)
    sanitized_url = mongo_url.split('@')[-1] if '@' in mongo_url else 'localhost'
    logger.info(f"✓ MONGO_URL configured (connecting to: {sanitized_url})")

# Configure MongoDB client with timeouts for production (Atlas)
# Note: Connection is lazy - actual connection happens on first operation
# Initialize client and db as None first
client = None
db = None

try:
    logger.info("Creating AsyncIOMotorClient...")
    client = AsyncIOMotorClient(
        mongo_url,
        serverSelectionTimeoutMS=10000,  # 10 seconds timeout for server selection
        connectTimeoutMS=20000,  # 20 seconds connection timeout
        socketTimeoutMS=45000,  # 45 seconds socket timeout
        maxPoolSize=50,  # Increased connection pool size
        minPoolSize=10,  # Minimum pool size
        retryWrites=True,  # Retry writes on network errors
        retryReads=True,  # Retry reads on network errors
    )
    db_name = os.environ.get('DB_NAME', 'aila_db')
    db = client[db_name]
    logger.info(f"✓ MongoDB client configured for database: {db_name}")
except Exception as e:
    logger.error(f"❌ Failed to configure MongoDB client: {e}", exc_info=True)
    # Don't raise - let the app start and health check will report the issue
    logger.warning("⚠ Continuing startup without MongoDB connection")
    # Create a dummy client that will fail gracefully
    try:
        logger.info("Creating fallback MongoDB client...")
        client = AsyncIOMotorClient("mongodb://localhost:27017", serverSelectionTimeoutMS=1000)
        db = client['aila_db']
        logger.info("✓ Fallback client created")
    except Exception as fallback_error:
        logger.error(f"❌ Fallback client creation failed: {fallback_error}")
        pass

# JWT Configuration - CRITIQUE POUR LA STABILITÉ DE L'AUTHENTIFICATION
# Ce secret DOIT être constant pour que les sessions restent valides après redémarrage
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    # Secret par défaut FIXE pour la production - NE JAMAIS CHANGER
    # Idéalement, définir JWT_SECRET dans les variables d'environnement Render
    JWT_SECRET = 'aila_production_jwt_secret_2024_stable_key_do_not_change_47e8cc1a'
    logging.warning("JWT_SECRET not set in environment, using default production secret")
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24 * 30  # 30 jours pour éviter les déconnexions fréquentes

# Create the main app
app = FastAPI(title="AÏLA - Arbre Généalogique API", version="1.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# ===================== MODELS =====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    first_name: str
    last_name: str
    gdpr_consent: bool = False

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class GoogleAuthRequest(BaseModel):
    """Request model for Google OAuth authentication"""
    id_token: Optional[str] = None  # Google ID token from frontend
    token: Optional[str] = None  # Alternative field name (from Google Identity Services)
    access_token: Optional[str] = None

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class UserResponse(BaseModel):
    id: str
    email: str
    first_name: str
    last_name: str
    created_at: datetime
    gdpr_consent: bool
    gdpr_consent_date: Optional[datetime] = None
    role: str = "member"  # admin or member

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class PersonCreate(BaseModel):
    first_name: str
    last_name: str
    gender: str = "unknown"  # male, female, unknown
    birth_date: Optional[str] = None
    birth_place: Optional[str] = None
    death_date: Optional[str] = None
    death_place: Optional[str] = None
    photo: Optional[str] = None  # base64
    notes: Optional[str] = None
    geographic_branch: Optional[str] = None  # région/zone géographique

class PersonUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    gender: Optional[str] = None
    birth_date: Optional[str] = None
    birth_place: Optional[str] = None
    death_date: Optional[str] = None
    death_place: Optional[str] = None
    photo: Optional[str] = None
    notes: Optional[str] = None
    geographic_branch: Optional[str] = None

class PersonResponse(BaseModel):
    id: str
    user_id: Optional[str] = None
    first_name: str
    last_name: str
    gender: str
    birth_date: Optional[str] = None
    birth_place: Optional[str] = None
    death_date: Optional[str] = None
    death_place: Optional[str] = None
    photo: Optional[str] = None
    notes: Optional[str] = None
    geographic_branch: Optional[str] = None
    created_at: datetime
    is_preview: bool = False

class FamilyLinkCreate(BaseModel):
    person_id_1: str
    person_id_2: str
    link_type: str  # parent, child, spouse

class FamilyLinkResponse(BaseModel):
    id: str
    user_id: Optional[str] = None
    person_id_1: str
    person_id_2: str
    link_type: str
    created_at: datetime

class PreviewSessionCreate(BaseModel):
    persons: List[PersonCreate] = []

class PreviewSessionResponse(BaseModel):
    session_token: str
    persons: List[PersonResponse]
    links: List[FamilyLinkResponse]
    created_at: datetime
    expires_at: datetime

class TreeResponse(BaseModel):
    persons: List[PersonResponse]
    links: List[FamilyLinkResponse]

# ===================== COLLABORATION MODELS =====================

class CollaboratorInvite(BaseModel):
    email: EmailStr
    role: str = "editor"  # viewer, editor

class CollaboratorResponse(BaseModel):
    id: str
    tree_owner_id: str
    user_id: Optional[str] = None
    email: str
    role: str
    status: str  # pending, accepted, rejected
    invited_at: datetime
    accepted_at: Optional[datetime] = None

class ContributionCreate(BaseModel):
    action: str  # add, edit, delete
    entity_type: str  # person, link
    entity_id: Optional[str] = None
    entity_data: Optional[dict] = None

class ContributionResponse(BaseModel):
    id: str
    tree_owner_id: str
    contributor_id: str
    contributor_name: str
    action: str
    entity_type: str
    entity_id: Optional[str] = None
    entity_data: Optional[dict] = None
    status: str  # pending, approved, rejected
    created_at: datetime
    reviewed_at: Optional[datetime] = None
    review_note: Optional[str] = None

class ContributionReview(BaseModel):
    status: str  # approved, rejected
    note: Optional[str] = None

class NotificationResponse(BaseModel):
    id: str
    user_id: str
    type: str
    title: str
    message: str
    data: Optional[dict] = None
    read: bool
    created_at: datetime

class ChatMessageCreate(BaseModel):
    message: str
    mentioned_person_id: Optional[str] = None  # ID d'une personne de l'arbre mentionnée

class ChatMessageResponse(BaseModel):
    id: str
    user_id: str
    user_name: str  # Prénom + Nom de l'expéditeur
    message: str
    mentioned_person_id: Optional[str] = None
    mentioned_person_name: Optional[str] = None
    created_at: datetime

# ===================== FAMILY EVENTS MODELS =====================

class FamilyEventCreate(BaseModel):
    event_type: str  # birthday, birth, graduation, wedding, newyear, holiday, custom
    title: str
    description: Optional[str] = None
    event_date: str  # ISO date string
    person_id: Optional[str] = None  # Related person if applicable
    recipients: List[str] = []  # List of collaborator emails to notify
    send_email: bool = False

class FamilyEventResponse(BaseModel):
    id: str
    user_id: str
    event_type: str
    title: str
    description: Optional[str] = None
    event_date: str
    person_id: Optional[str] = None
    person_name: Optional[str] = None
    recipients: List[str] = []
    send_email: bool
    created_at: datetime

class UpcomingBirthdayResponse(BaseModel):
    person_id: str
    person_name: str
    birth_date: str
    days_until: int
    age: Optional[int] = None

# ===================== TREE MERGE MODELS =====================

class DuplicateCandidate(BaseModel):
    source_person_id: str
    source_person_name: str
    source_birth_date: Optional[str] = None
    target_person_id: str
    target_person_name: str
    target_birth_date: Optional[str] = None
    similarity_score: float  # 0-100
    match_reason: str

class MergeAnalysisResponse(BaseModel):
    source_tree_owner_id: str
    source_tree_owner_name: str
    source_persons_count: int
    target_persons_count: int
    duplicates_found: List[DuplicateCandidate]
    new_persons_count: int  # Persons that will be added (not duplicates)
    new_links_count: int

class MergeDecision(BaseModel):
    source_person_id: str
    action: str  # "merge" (use target), "add" (add as new), "skip" (ignore)
    target_person_id: Optional[str] = None  # Required if action is "merge"

class MergeExecuteRequest(BaseModel):
    source_tree_owner_id: str
    decisions: List[MergeDecision]
    import_links: bool = True

class MergeExecuteResponse(BaseModel):
    persons_merged: int
    persons_added: int
    persons_skipped: int
    links_added: int
    message: str

# ===================== HELPERS =====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str) -> str:
    payload = {
        'user_id': user_id,
        'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS),
        'iat': datetime.utcnow()
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    payload = decode_token(credentials.credentials)
    user = await db.users.find_one({"_id": ObjectId(payload['user_id'])})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

def serialize_object_id(obj: dict) -> dict:
    if obj and '_id' in obj:
        obj['id'] = str(obj['_id'])
        del obj['_id']
    return obj

# ===================== AUTH ROUTES =====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    # Check if email exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    if not user_data.gdpr_consent:
        raise HTTPException(status_code=400, detail="GDPR consent is required")
    
    # Check if this is the first user - make them admin
    user_count = await db.users.count_documents({})
    role = "admin" if user_count == 0 else "member"
    
    # Create user
    user_doc = {
        "email": user_data.email,
        "password_hash": hash_password(user_data.password),
        "first_name": user_data.first_name,
        "last_name": user_data.last_name,
        "gdpr_consent": user_data.gdpr_consent,
        "gdpr_consent_date": datetime.utcnow() if user_data.gdpr_consent else None,
        "created_at": datetime.utcnow(),
        "role": role
    }
    
    result = await db.users.insert_one(user_doc)
    user_doc['_id'] = result.inserted_id
    user_doc = serialize_object_id(user_doc)
    
    token = create_token(user_doc['id'])
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user_doc['id'],
            email=user_doc['email'],
            first_name=user_doc['first_name'],
            last_name=user_doc['last_name'],
            created_at=user_doc['created_at'],
            gdpr_consent=user_doc['gdpr_consent'],
            gdpr_consent_date=user_doc.get('gdpr_consent_date'),
            role=user_doc['role']
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email})
    if not user or not verify_password(credentials.password, user['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    user = serialize_object_id(user)
    token = create_token(user['id'])
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user['id'],
            email=user['email'],
            first_name=user['first_name'],
            last_name=user['last_name'],
            created_at=user['created_at'],
            gdpr_consent=user.get('gdpr_consent', False),
            gdpr_consent_date=user.get('gdpr_consent_date'),
            role=user.get('role', 'member')
        )
    )

@api_router.post("/auth/google", response_model=TokenResponse)
async def google_auth(data: GoogleAuthRequest):
    """Authenticate or register user via Google OAuth"""
    import httpx
    
    # Accept either token or id_token field
    google_token = data.token or data.id_token
    if not google_token:
        raise HTTPException(status_code=400, detail="No Google token provided")
    
    try:
        # Verify the Google ID token
        async with httpx.AsyncClient() as client:
            response = await client.get(
                f"https://oauth2.googleapis.com/tokeninfo?id_token={google_token}"
            )
            
            if response.status_code != 200:
                raise HTTPException(status_code=401, detail="Invalid Google token")
            
            google_data = response.json()
            
            # Verify the token is for our app (optional - add your client ID)
            # if google_data.get('aud') != GOOGLE_CLIENT_ID:
            #     raise HTTPException(status_code=401, detail="Token not for this app")
            
            email = google_data.get('email', '').lower().strip()
            if not email:
                raise HTTPException(status_code=400, detail="Email not provided by Google")
            
            # Check if user exists
            user = await db.users.find_one({"email": email})
            
            if user:
                # Existing user - login
                user = serialize_object_id(user)
                token = create_token(user['id'])
                logger.info(f"✓ Google login successful for {email}")
            else:
                # New user - register
                first_name = google_data.get('given_name', google_data.get('name', 'Utilisateur'))
                last_name = google_data.get('family_name', '')
                
                # Generate a random password (user won't need it for Google login)
                import secrets
                random_password = secrets.token_urlsafe(32)
                
                new_user = {
                    "email": email,
                    "password_hash": hash_password(random_password),
                    "first_name": first_name,
                    "last_name": last_name,
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow(),
                    "gdpr_consent": True,
                    "gdpr_consent_date": datetime.utcnow(),
                    "role": "member",
                    "auth_provider": "google",
                    "google_id": google_data.get('sub'),
                    "profile_picture": google_data.get('picture')
                }
                
                result = await db.users.insert_one(new_user)
                new_user['id'] = str(result.inserted_id)
                user = new_user
                token = create_token(user['id'])
                logger.info(f"✓ New Google user registered: {email}")
            
            return TokenResponse(
                access_token=token,
                user=UserResponse(
                    id=user['id'],
                    email=user['email'],
                    first_name=user['first_name'],
                    last_name=user['last_name'],
                    created_at=user.get('created_at', datetime.utcnow()),
                    gdpr_consent=user.get('gdpr_consent', True),
                    gdpr_consent_date=user.get('gdpr_consent_date'),
                    role=user.get('role', 'member')
                )
            )
            
    except httpx.HTTPError as e:
        logger.error(f"Google auth HTTP error: {e}")
        raise HTTPException(status_code=500, detail="Failed to verify Google token")
    except Exception as e:
        logger.error(f"Google auth error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# GOOGLE OAUTH SERVER-SIDE FLOW
# ============================================================================

@api_router.get("/auth/google/login")
async def google_login_redirect():
    """Redirect user to Google OAuth consent screen"""
    import urllib.parse
    
    # Build Google OAuth URL
    params = {
        'client_id': GOOGLE_CLIENT_ID,
        'redirect_uri': f"{FRONTEND_URL}/api/auth/google/callback",
        'response_type': 'code',
        'scope': 'openid email profile',
        'access_type': 'offline',
        'prompt': 'select_account',
    }
    
    google_auth_url = f"https://accounts.google.com/o/oauth2/v2/auth?{urllib.parse.urlencode(params)}"
    return RedirectResponse(url=google_auth_url)

@api_router.get("/auth/google/callback")
async def google_callback(code: str = None, error: str = None):
    """Handle Google OAuth callback"""
    if error:
        logger.error(f"Google OAuth error: {error}")
        return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=google_auth_failed")
    
    if not code:
        return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=no_code")
    
    try:
        # Log pour déboguer
        logger.info(f"Google callback - exchanging code for token...")
        logger.info(f"Using client_id: {GOOGLE_CLIENT_ID[:20]}...")
        logger.info(f"Using client_secret: {'SET' if GOOGLE_CLIENT_SECRET else 'NOT SET'} ({len(GOOGLE_CLIENT_SECRET)} chars)")
        logger.info(f"Using redirect_uri: {FRONTEND_URL}/api/auth/google/callback")
        
        # Exchange code for tokens
        async with httpx.AsyncClient() as client:
            token_response = await client.post(
                "https://oauth2.googleapis.com/token",
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                data={
                    'client_id': GOOGLE_CLIENT_ID,
                    'client_secret': GOOGLE_CLIENT_SECRET,
                    'code': code,
                    'grant_type': 'authorization_code',
                    'redirect_uri': f"{FRONTEND_URL}/api/auth/google/callback",
                }
            )
            
            if token_response.status_code != 200:
                logger.error(f"Google token exchange failed: {token_response.text}")
                return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=token_exchange_failed")
            
            tokens = token_response.json()
            id_token = tokens.get('id_token')
            
            if not id_token:
                return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=no_id_token")
            
            # Verify and decode the ID token
            token_info_response = await client.get(
                f"https://oauth2.googleapis.com/tokeninfo?id_token={id_token}"
            )
            
            if token_info_response.status_code != 200:
                return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=invalid_token")
            
            google_data = token_info_response.json()
            email = google_data.get('email', '').lower().strip()
            
            if not email:
                return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=no_email")
            
            # Check if user exists
            user = await db.users.find_one({"email": email})
            
            if user:
                # Existing user - login
                user = serialize_object_id(user)
                token = create_token(user['id'])
                logger.info(f"✓ Google OAuth login successful for {email}")
            else:
                # New user - register
                import secrets
                first_name = google_data.get('given_name', google_data.get('name', 'Utilisateur'))
                last_name = google_data.get('family_name', '')
                random_password = secrets.token_urlsafe(32)
                
                new_user = {
                    "email": email,
                    "password_hash": hash_password(random_password),
                    "first_name": first_name,
                    "last_name": last_name,
                    "created_at": datetime.utcnow(),
                    "updated_at": datetime.utcnow(),
                    "gdpr_consent": True,
                    "gdpr_consent_date": datetime.utcnow(),
                    "role": "member",
                    "auth_provider": "google",
                    "google_id": google_data.get('sub'),
                    "profile_picture": google_data.get('picture')
                }
                
                result = await db.users.insert_one(new_user)
                new_user['id'] = str(result.inserted_id)
                user = new_user
                token = create_token(user['id'])
                logger.info(f"✓ New Google OAuth user registered: {email}")
            
            # Redirect to frontend with token
            return RedirectResponse(
                url=f"{FRONTEND_URL}/(tabs)/tree?token={token}&google_auth=success"
            )
            
    except Exception as e:
        logger.error(f"Google OAuth callback error: {e}")
        return RedirectResponse(url=f"{FRONTEND_URL}/(auth)/login?error=server_error")

# Temporary admin endpoint to reset password - REMOVE AFTER USE
class PasswordReset(BaseModel):
    email: EmailStr
    new_password: str
    admin_key: str

@api_router.post("/auth/admin-reset-password")
async def admin_reset_password(data: PasswordReset):
    # Simple admin key protection
    if data.admin_key != "AILA_TEMP_RESET_2024":
        raise HTTPException(status_code=403, detail="Invalid admin key")
    
    user = await db.users.find_one({"email": data.email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_hash = hash_password(data.new_password)
    await db.users.update_one(
        {"email": data.email},
        {"$set": {"password_hash": new_hash}}
    )
    
    return {"message": "Password reset successfully"}

@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPasswordRequest):
    """Request a password reset email"""
    email = data.email.lower().strip()
    
    # Check if user exists (but don't reveal if they don't for security)
    user = await db.users.find_one({"email": {"$regex": f"^{email}$", "$options": "i"}})
    
    if user:
        # Generate reset token
        reset_token = str(uuid.uuid4())
        expires_at = datetime.utcnow() + timedelta(hours=1)
        
        # Store reset token in database
        await db.password_resets.delete_many({"email": email})  # Remove old tokens
        await db.password_resets.insert_one({
            "email": email,
            "token": reset_token,
            "expires_at": expires_at,
            "created_at": datetime.utcnow()
        })
        
        # Send email
        await send_password_reset_email(email, reset_token)
    
    # Always return success to prevent email enumeration
    return {"message": "Si un compte existe avec cet email, vous recevrez un lien de réinitialisation."}

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordRequest):
    """Reset password using the token from email"""
    # Find the reset token
    reset_record = await db.password_resets.find_one({"token": data.token})
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Lien de réinitialisation invalide ou expiré")
    
    # Check if token is expired
    if reset_record['expires_at'] < datetime.utcnow():
        await db.password_resets.delete_one({"token": data.token})
        raise HTTPException(status_code=400, detail="Ce lien a expiré. Veuillez demander un nouveau lien.")
    
    # Validate new password
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Le mot de passe doit contenir au moins 6 caractères")
    
    # Update password
    new_hash = hash_password(data.new_password)
    result = await db.users.update_one(
        {"email": {"$regex": f"^{reset_record['email']}$", "$options": "i"}},
        {"$set": {"password_hash": new_hash}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Delete the used token
    await db.password_resets.delete_one({"token": data.token})
    
    return {"message": "Votre mot de passe a été réinitialisé avec succès. Vous pouvez maintenant vous connecter."}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    user = serialize_object_id(current_user.copy())
    return UserResponse(
        id=user['id'],
        email=user['email'],
        first_name=user['first_name'],
        last_name=user['last_name'],
        created_at=user['created_at'],
        gdpr_consent=user.get('gdpr_consent', False),
        gdpr_consent_date=user.get('gdpr_consent_date'),
        role=user.get('role', 'member')
    )

# ===================== USER MANAGEMENT ROUTES (ADMIN ONLY) =====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_all_users(current_user: dict = Depends(get_current_user)):
    # Only admins can list all users
    if current_user.get('role', 'member') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}).to_list(500)
    return [
        UserResponse(
            id=str(u['_id']),
            email=u['email'],
            first_name=u['first_name'],
            last_name=u['last_name'],
            created_at=u['created_at'],
            gdpr_consent=u.get('gdpr_consent', False),
            gdpr_consent_date=u.get('gdpr_consent_date'),
            role=u.get('role', 'member')
        )
        for u in users
    ]

@api_router.put("/users/{user_id}/promote")
async def promote_user_to_admin(user_id: str, current_user: dict = Depends(get_current_user)):
    # Only admins can promote users
    if current_user.get('role', 'member') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if target user exists
    target_user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update user role
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"role": "admin"}}
    )
    
    return {"message": f"User {target_user['email']} promoted to admin"}

@api_router.put("/users/{user_id}/demote")
async def demote_user_to_member(user_id: str, current_user: dict = Depends(get_current_user)):
    # Only admins can demote users
    if current_user.get('role', 'member') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Prevent demoting yourself
    if str(current_user['_id']) == user_id:
        raise HTTPException(status_code=400, detail="Cannot demote yourself")
    
    # Check if target user exists
    target_user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update user role
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"role": "member"}}
    )
    
    return {"message": f"User {target_user['email']} demoted to member"}


# ===================== PERSON ROUTES =====================

@api_router.post("/persons", response_model=PersonResponse)
async def create_person(person: PersonCreate, current_user: dict = Depends(get_current_user)):
    person_doc = {
        "user_id": str(current_user['_id']),
        "first_name": person.first_name,
        "last_name": person.last_name,
        "gender": person.gender,
        "birth_date": person.birth_date,
        "birth_place": person.birth_place,
        "death_date": person.death_date,
        "death_place": person.death_place,
        "photo": person.photo,
        "notes": person.notes,
        "geographic_branch": person.geographic_branch,
        "created_at": datetime.utcnow(),
        "is_preview": False
    }
    
    result = await db.persons.insert_one(person_doc)
    person_doc['_id'] = result.inserted_id
    person_doc = serialize_object_id(person_doc)
    
    return PersonResponse(**person_doc)

@api_router.get("/persons", response_model=List[PersonResponse])
async def get_persons(current_user: dict = Depends(get_current_user)):
    persons = await db.persons.find({"user_id": str(current_user['_id'])}).to_list(500)
    return [PersonResponse(**serialize_object_id(p)) for p in persons]

@api_router.get("/persons/{person_id}", response_model=PersonResponse)
async def get_person(person_id: str, current_user: dict = Depends(get_current_user)):
    person = await db.persons.find_one({
        "_id": ObjectId(person_id),
        "user_id": str(current_user['_id'])
    })
    if not person:
        raise HTTPException(status_code=404, detail="Person not found")
    return PersonResponse(**serialize_object_id(person))

@api_router.put("/persons/{person_id}", response_model=PersonResponse)
async def update_person(person_id: str, person: PersonUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in person.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # Check if person exists
    existing_person = await db.persons.find_one({"_id": ObjectId(person_id)})
    if not existing_person:
        raise HTTPException(status_code=404, detail="Person not found")
    
    # Check permissions: admin can edit anything, members can only edit their own
    user_role = current_user.get('role', 'member')
    if user_role != 'admin' and existing_person['user_id'] != str(current_user['_id']):
        raise HTTPException(status_code=403, detail="You can only modify your own entries")
    
    result = await db.persons.find_one_and_update(
        {"_id": ObjectId(person_id)},
        {"$set": update_data},
        return_document=True
    )
    return PersonResponse(**serialize_object_id(result))

@api_router.delete("/persons/{person_id}")
async def delete_person(person_id: str, current_user: dict = Depends(get_current_user)):
    # Check if person exists
    existing_person = await db.persons.find_one({"_id": ObjectId(person_id)})
    if not existing_person:
        raise HTTPException(status_code=404, detail="Person not found")
    
    # Check permissions: admin can delete anything, members can only delete their own
    user_role = current_user.get('role', 'member')
    if user_role != 'admin' and existing_person['user_id'] != str(current_user['_id']):
        raise HTTPException(status_code=403, detail="You can only delete your own entries")
    
    result = await db.persons.delete_one({"_id": ObjectId(person_id)})
    
    # Also delete related links
    await db.family_links.delete_many({
        "$or": [
            {"person_id_1": person_id},
            {"person_id_2": person_id}
        ]
    })
    
    return {"message": "Person deleted successfully"}

# ===================== FAMILY LINK ROUTES =====================

@api_router.post("/links", response_model=FamilyLinkResponse)
async def create_link(link: FamilyLinkCreate, current_user: dict = Depends(get_current_user)):
    # Verify both persons exist and belong to user
    person1 = await db.persons.find_one({
        "_id": ObjectId(link.person_id_1),
        "user_id": str(current_user['_id'])
    })
    person2 = await db.persons.find_one({
        "_id": ObjectId(link.person_id_2),
        "user_id": str(current_user['_id'])
    })
    
    if not person1 or not person2:
        raise HTTPException(status_code=404, detail="One or both persons not found")
    
    # Extended family relationship types
    valid_link_types = [
        'parent', 'child', 'spouse',  # Core relationships
        'sibling',  # Frère/Sœur
        'grandparent', 'grandchild',  # Grand-parent / Petit-enfant
        'uncle_aunt', 'nephew_niece',  # Oncle-Tante / Neveu-Nièce
        'cousin',  # Cousin/Cousine
        'step_parent', 'step_child',  # Beau-parent / Beau-fils
        'parent_in_law', 'child_in_law',  # Beau-père-mère / Gendre-Belle-fille
        'sibling_in_law',  # Beau-frère / Belle-sœur
        'godparent', 'godchild',  # Parrain-Marraine / Filleul(e)
    ]
    
    if link.link_type not in valid_link_types:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid link type. Valid types: {', '.join(valid_link_types)}"
        )
    
    link_doc = {
        "user_id": str(current_user['_id']),
        "person_id_1": link.person_id_1,
        "person_id_2": link.person_id_2,
        "link_type": link.link_type,
        "created_at": datetime.utcnow()
    }
    
    result = await db.family_links.insert_one(link_doc)
    link_doc['_id'] = result.inserted_id
    link_doc = serialize_object_id(link_doc)
    
    return FamilyLinkResponse(**link_doc)

@api_router.get("/links", response_model=List[FamilyLinkResponse])
async def get_links(current_user: dict = Depends(get_current_user)):
    links = await db.family_links.find({"user_id": str(current_user['_id'])}).to_list(500)
    return [FamilyLinkResponse(**serialize_object_id(l)) for l in links]

@api_router.delete("/links/{link_id}")
async def delete_link(link_id: str, current_user: dict = Depends(get_current_user)):
    # Check if link exists
    existing_link = await db.family_links.find_one({"_id": ObjectId(link_id)})
    if not existing_link:
        raise HTTPException(status_code=404, detail="Link not found")
    
    # Check permissions: admin can delete anything, members can only delete their own
    user_role = current_user.get('role', 'member')
    if user_role != 'admin' and existing_link['user_id'] != str(current_user['_id']):
        raise HTTPException(status_code=403, detail="You can only delete your own entries")
    
    result = await db.family_links.delete_one({"_id": ObjectId(link_id)})
    return {"message": "Link deleted successfully"}

# ===================== TREE ROUTES =====================

@api_router.get("/tree", response_model=TreeResponse)
async def get_tree(current_user: dict = Depends(get_current_user)):
    persons = await db.persons.find({"user_id": str(current_user['_id'])}).to_list(500)
    links = await db.family_links.find({"user_id": str(current_user['_id'])}).to_list(500)
    
    return TreeResponse(
        persons=[PersonResponse(**serialize_object_id(p)) for p in persons],
        links=[FamilyLinkResponse(**serialize_object_id(l)) for l in links]
    )

# ===================== PREVIEW MODE ROUTES =====================

@api_router.post("/preview/session", response_model=PreviewSessionResponse)
async def create_preview_session():
    session_token = str(uuid.uuid4())
    expires_at = datetime.utcnow() + timedelta(hours=24)  # 24h expiration
    
    session_doc = {
        "session_token": session_token,
        "persons": [],
        "links": [],
        "created_at": datetime.utcnow(),
        "expires_at": expires_at
    }
    
    await db.preview_sessions.insert_one(session_doc)
    
    return PreviewSessionResponse(
        session_token=session_token,
        persons=[],
        links=[],
        created_at=session_doc['created_at'],
        expires_at=expires_at
    )

@api_router.post("/preview/demo", response_model=PreviewSessionResponse)
async def create_demo_session():
    """
    Crée une session de démonstration avec une famille exemple pré-remplie.
    Permet aux nouveaux utilisateurs de voir immédiatement la valeur de l'application.
    """
    session_token = str(uuid.uuid4())
    expires_at = datetime.utcnow() + timedelta(hours=24)
    now = datetime.utcnow()
    
    # Famille Martin - Exemple réaliste sur 3 générations
    # IDs uniques pour chaque personne
    gp_henri_id = str(uuid.uuid4())
    gp_marie_id = str(uuid.uuid4())
    gp_jean_id = str(uuid.uuid4())
    gp_jeanne_id = str(uuid.uuid4())
    
    p_pierre_id = str(uuid.uuid4())
    p_sophie_id = str(uuid.uuid4())
    
    c_lucas_id = str(uuid.uuid4())
    c_emma_id = str(uuid.uuid4())
    c_leo_id = str(uuid.uuid4())
    
    demo_persons = [
        # Génération 1 - Grands-parents paternels
        {
            "id": gp_henri_id,
            "first_name": "Henri",
            "last_name": "Martin",
            "gender": "male",
            "birth_date": "1945-03-15",
            "birth_place": "Lyon",
            "notes": "Grand-père paternel - Ancien instituteur",
            "created_at": now
        },
        {
            "id": gp_marie_id,
            "first_name": "Marie",
            "last_name": "Martin",
            "gender": "female",
            "birth_date": "1948-07-22",
            "birth_place": "Lyon",
            "notes": "Grand-mère paternelle - Née Dubois",
            "created_at": now
        },
        # Génération 1 - Grands-parents maternels
        {
            "id": gp_jean_id,
            "first_name": "Jean",
            "last_name": "Bernard",
            "gender": "male",
            "birth_date": "1942-11-08",
            "birth_place": "Paris",
            "death_date": "2020-04-12",
            "notes": "Grand-père maternel - Ingénieur",
            "created_at": now
        },
        {
            "id": gp_jeanne_id,
            "first_name": "Jeanne",
            "last_name": "Bernard",
            "gender": "female",
            "birth_date": "1946-02-28",
            "birth_place": "Paris",
            "notes": "Grand-mère maternelle - Née Petit",
            "created_at": now
        },
        # Génération 2 - Parents
        {
            "id": p_pierre_id,
            "first_name": "Pierre",
            "last_name": "Martin",
            "gender": "male",
            "birth_date": "1975-06-10",
            "birth_place": "Lyon",
            "notes": "Père - Médecin généraliste",
            "created_at": now
        },
        {
            "id": p_sophie_id,
            "first_name": "Sophie",
            "last_name": "Martin",
            "gender": "female",
            "birth_date": "1978-09-03",
            "birth_place": "Paris",
            "notes": "Mère - Avocate - Née Bernard",
            "created_at": now
        },
        # Génération 3 - Enfants
        {
            "id": c_lucas_id,
            "first_name": "Lucas",
            "last_name": "Martin",
            "gender": "male",
            "birth_date": "2005-01-20",
            "birth_place": "Lyon",
            "notes": "Fils aîné - Étudiant",
            "created_at": now
        },
        {
            "id": c_emma_id,
            "first_name": "Emma",
            "last_name": "Martin",
            "gender": "female",
            "birth_date": "2008-04-14",
            "birth_place": "Lyon",
            "notes": "Fille - Lycéenne",
            "created_at": now
        },
        {
            "id": c_leo_id,
            "first_name": "Léo",
            "last_name": "Martin",
            "gender": "male",
            "birth_date": "2012-08-30",
            "birth_place": "Lyon",
            "notes": "Fils cadet - Collégien",
            "created_at": now
        },
    ]
    
    demo_links = [
        # Couples
        {"id": str(uuid.uuid4()), "person_id_1": gp_henri_id, "person_id_2": gp_marie_id, "link_type": "spouse", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": gp_jean_id, "person_id_2": gp_jeanne_id, "link_type": "spouse", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": p_pierre_id, "person_id_2": p_sophie_id, "link_type": "spouse", "created_at": now},
        
        # Henri & Marie -> Pierre
        {"id": str(uuid.uuid4()), "person_id_1": gp_henri_id, "person_id_2": p_pierre_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": gp_marie_id, "person_id_2": p_pierre_id, "link_type": "parent", "created_at": now},
        
        # Jean & Jeanne -> Sophie
        {"id": str(uuid.uuid4()), "person_id_1": gp_jean_id, "person_id_2": p_sophie_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": gp_jeanne_id, "person_id_2": p_sophie_id, "link_type": "parent", "created_at": now},
        
        # Pierre & Sophie -> Lucas, Emma, Léo
        {"id": str(uuid.uuid4()), "person_id_1": p_pierre_id, "person_id_2": c_lucas_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": p_sophie_id, "person_id_2": c_lucas_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": p_pierre_id, "person_id_2": c_emma_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": p_sophie_id, "person_id_2": c_emma_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": p_pierre_id, "person_id_2": c_leo_id, "link_type": "parent", "created_at": now},
        {"id": str(uuid.uuid4()), "person_id_1": p_sophie_id, "person_id_2": c_leo_id, "link_type": "parent", "created_at": now},
    ]
    
    session_doc = {
        "session_token": session_token,
        "persons": demo_persons,
        "links": demo_links,
        "created_at": now,
        "expires_at": expires_at,
        "is_demo": True
    }
    
    await db.preview_sessions.insert_one(session_doc)
    
    # Convert to response format
    persons_response = [PersonResponse(
        id=p['id'],
        first_name=p['first_name'],
        last_name=p['last_name'],
        gender=p.get('gender', 'unknown'),
        birth_date=p.get('birth_date'),
        birth_place=p.get('birth_place'),
        death_date=p.get('death_date'),
        death_place=p.get('death_place'),
        photo=p.get('photo'),
        notes=p.get('notes'),
        geographic_branch=p.get('geographic_branch'),
        created_at=p.get('created_at', now),
        is_preview=True
    ) for p in demo_persons]
    
    links_response = [FamilyLinkResponse(
        id=l['id'],
        person_id_1=l['person_id_1'],
        person_id_2=l['person_id_2'],
        link_type=l['link_type'],
        created_at=l.get('created_at', now)
    ) for l in demo_links]
    
    return PreviewSessionResponse(
        session_token=session_token,
        persons=persons_response,
        links=links_response,
        created_at=now,
        expires_at=expires_at
    )

@api_router.get("/preview/{session_token}", response_model=PreviewSessionResponse)
async def get_preview_session(session_token: str):
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    if session['expires_at'] < datetime.utcnow():
        await db.preview_sessions.delete_one({"session_token": session_token})
        raise HTTPException(status_code=410, detail="Preview session expired")
    
    persons = [PersonResponse(
        id=p.get('id', str(uuid.uuid4())),
        first_name=p['first_name'],
        last_name=p['last_name'],
        gender=p.get('gender', 'unknown'),
        birth_date=p.get('birth_date'),
        birth_place=p.get('birth_place'),
        death_date=p.get('death_date'),
        death_place=p.get('death_place'),
        photo=p.get('photo'),
        notes=p.get('notes'),
        geographic_branch=p.get('geographic_branch'),
        created_at=p.get('created_at', datetime.utcnow()),
        is_preview=True
    ) for p in session.get('persons', [])]
    
    links = [FamilyLinkResponse(
        id=l.get('id', str(uuid.uuid4())),
        person_id_1=l['person_id_1'],
        person_id_2=l['person_id_2'],
        link_type=l['link_type'],
        created_at=l.get('created_at', datetime.utcnow())
    ) for l in session.get('links', [])]
    
    return PreviewSessionResponse(
        session_token=session_token,
        persons=persons,
        links=links,
        created_at=session['created_at'],
        expires_at=session['expires_at']
    )

@api_router.post("/preview/{session_token}/person", response_model=PreviewSessionResponse)
async def add_preview_person(session_token: str, person: PersonCreate):
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    if len(session.get('persons', [])) >= 10:
        raise HTTPException(status_code=400, detail="Preview mode limited to 10 members")
    
    person_doc = {
        "id": str(uuid.uuid4()),
        "first_name": person.first_name,
        "last_name": person.last_name,
        "gender": person.gender,
        "birth_date": person.birth_date,
        "birth_place": person.birth_place,
        "death_date": person.death_date,
        "death_place": person.death_place,
        "photo": person.photo,
        "notes": person.notes,
        "geographic_branch": person.geographic_branch,
        "created_at": datetime.utcnow().isoformat()
    }
    
    await db.preview_sessions.update_one(
        {"session_token": session_token},
        {"$push": {"persons": person_doc}}
    )
    
    return await get_preview_session(session_token)

@api_router.post("/preview/{session_token}/link", response_model=PreviewSessionResponse)
async def add_preview_link(session_token: str, link: FamilyLinkCreate):
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    link_doc = {
        "id": str(uuid.uuid4()),
        "person_id_1": link.person_id_1,
        "person_id_2": link.person_id_2,
        "link_type": link.link_type,
        "created_at": datetime.utcnow().isoformat()
    }
    
    await db.preview_sessions.update_one(
        {"session_token": session_token},
        {"$push": {"links": link_doc}}
    )
    
    return await get_preview_session(session_token)

@api_router.delete("/preview/{session_token}/person/{person_id}", response_model=PreviewSessionResponse)
async def delete_preview_person(session_token: str, person_id: str):
    """Delete a person from preview session"""
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    if session['expires_at'] < datetime.utcnow():
        await db.preview_sessions.delete_one({"session_token": session_token})
        raise HTTPException(status_code=410, detail="Preview session expired")
    
    # Remove person from persons list
    persons = session.get('persons', [])
    original_count = len(persons)
    persons = [p for p in persons if p.get('id') != person_id]
    
    if len(persons) == original_count:
        raise HTTPException(status_code=404, detail="Person not found in preview session")
    
    # Also remove any links involving this person
    links = session.get('links', [])
    links = [l for l in links if l.get('person_id_1') != person_id and l.get('person_id_2') != person_id]
    
    await db.preview_sessions.update_one(
        {"session_token": session_token},
        {"$set": {"persons": persons, "links": links}}
    )
    
    return await get_preview_session(session_token)

@api_router.put("/preview/{session_token}/person/{person_id}", response_model=PreviewSessionResponse)
async def update_preview_person(session_token: str, person_id: str, person: PersonCreate):
    """Update a person in preview session"""
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    if session['expires_at'] < datetime.utcnow():
        await db.preview_sessions.delete_one({"session_token": session_token})
        raise HTTPException(status_code=410, detail="Preview session expired")
    
    # Find and update person
    persons = session.get('persons', [])
    person_found = False
    for i, p in enumerate(persons):
        if p.get('id') == person_id:
            persons[i] = {
                "id": person_id,
                "first_name": person.first_name,
                "last_name": person.last_name,
                "gender": person.gender,
                "birth_date": person.birth_date,
                "birth_place": person.birth_place,
                "death_date": person.death_date,
                "death_place": person.death_place,
                "geographic_branch": person.geographic_branch,
                "notes": person.notes,
            }
            person_found = True
            break
    
    if not person_found:
        raise HTTPException(status_code=404, detail="Person not found in preview session")
    
    await db.preview_sessions.update_one(
        {"session_token": session_token},
        {"$set": {"persons": persons}}
    )
    
    return await get_preview_session(session_token)

@api_router.delete("/preview/{session_token}/link/{link_id}", response_model=PreviewSessionResponse)
async def delete_preview_link(session_token: str, link_id: str):
    """Delete a link from preview session"""
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    if session['expires_at'] < datetime.utcnow():
        await db.preview_sessions.delete_one({"session_token": session_token})
        raise HTTPException(status_code=410, detail="Preview session expired")
    
    # Remove link
    links = session.get('links', [])
    original_count = len(links)
    links = [l for l in links if l.get('id') != link_id]
    
    if len(links) == original_count:
        raise HTTPException(status_code=404, detail="Link not found in preview session")
    
    await db.preview_sessions.update_one(
        {"session_token": session_token},
        {"$set": {"links": links}}
    )
    
    return await get_preview_session(session_token)

@api_router.post("/preview/{session_token}/convert")
async def convert_preview_to_permanent(session_token: str, current_user: dict = Depends(get_current_user)):
    session = await db.preview_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=404, detail="Preview session not found")
    
    user_id = str(current_user['_id'])
    old_to_new_ids = {}
    
    # Convert persons
    for p in session.get('persons', []):
        old_id = p.get('id')
        person_doc = {
            "user_id": user_id,
            "first_name": p['first_name'],
            "last_name": p['last_name'],
            "gender": p.get('gender', 'unknown'),
            "birth_date": p.get('birth_date'),
            "birth_place": p.get('birth_place'),
            "death_date": p.get('death_date'),
            "death_place": p.get('death_place'),
            "photo": p.get('photo'),
            "notes": p.get('notes'),
            "geographic_branch": p.get('geographic_branch'),
            "created_at": datetime.utcnow(),
            "is_preview": False
        }
        result = await db.persons.insert_one(person_doc)
        old_to_new_ids[old_id] = str(result.inserted_id)
    
    # Convert links with updated IDs
    for l in session.get('links', []):
        new_person_1 = old_to_new_ids.get(l['person_id_1'])
        new_person_2 = old_to_new_ids.get(l['person_id_2'])
        if new_person_1 and new_person_2:
            link_doc = {
                "user_id": user_id,
                "person_id_1": new_person_1,
                "person_id_2": new_person_2,
                "link_type": l['link_type'],
                "created_at": datetime.utcnow()
            }
            await db.family_links.insert_one(link_doc)
    
    # Delete preview session
    await db.preview_sessions.delete_one({"session_token": session_token})
    
    return {"message": "Preview converted to permanent tree", "persons_converted": len(old_to_new_ids)}

# ===================== GDPR ROUTES =====================

@api_router.get("/gdpr/export")
async def export_user_data(current_user: dict = Depends(get_current_user)):
    """Export all user data for GDPR compliance"""
    user_id = str(current_user['_id'])
    
    persons = await db.persons.find({"user_id": user_id}).to_list(500)
    links = await db.family_links.find({"user_id": user_id}).to_list(500)
    
    user_data = serialize_object_id(current_user.copy())
    if 'password_hash' in user_data:
        del user_data['password_hash']
    
    return {
        "user": user_data,
        "persons": [serialize_object_id(p) for p in persons],
        "family_links": [serialize_object_id(l) for l in links],
        "exported_at": datetime.utcnow().isoformat()
    }

@api_router.delete("/gdpr/delete-account")
async def delete_account(current_user: dict = Depends(get_current_user)):
    """Delete all user data for GDPR compliance"""
    user_id = str(current_user['_id'])
    
    await db.persons.delete_many({"user_id": user_id})
    await db.family_links.delete_many({"user_id": user_id})
    await db.collaborators.delete_many({"$or": [{"tree_owner_id": user_id}, {"user_id": user_id}]})
    await db.contributions.delete_many({"$or": [{"tree_owner_id": user_id}, {"contributor_id": user_id}]})
    await db.notifications.delete_many({"user_id": user_id})
    await db.users.delete_one({"_id": current_user['_id']})
    
    return {"message": "Account and all data deleted successfully"}

# ===================== TREE MANAGEMENT ROUTES =====================

@api_router.delete("/tree/clear")
async def clear_tree(current_user: dict = Depends(get_current_user)):
    """Delete all persons and links in the user's tree (keeps account)"""
    user_id = str(current_user['_id'])
    
    # Delete all persons and links
    deleted_persons = await db.persons.delete_many({"user_id": user_id})
    deleted_links = await db.family_links.delete_many({"user_id": user_id})
    
    return {
        "message": "Arbre supprimé avec succès",
        "deleted_persons": deleted_persons.deleted_count,
        "deleted_links": deleted_links.deleted_count
    }

# ===================== DEBUG ROUTES =====================

@api_router.get("/tree/debug")
async def debug_tree(current_user: dict = Depends(get_current_user)):
    """Get detailed tree data for debugging"""
    user_id = str(current_user['_id'])
    
    persons = await db.persons.find({"user_id": user_id}).to_list(500)
    links = await db.family_links.find({"user_id": user_id}).to_list(500)
    
    # Build person map
    person_map = {str(p['_id']): f"{p['first_name']} {p['last_name']}" for p in persons}
    
    # Annotate links with names
    annotated_links = []
    for link in links:
        annotated_links.append({
            "id": str(link['_id']),
            "type": link['link_type'],
            "person_1": {
                "id": link['person_id_1'],
                "name": person_map.get(link['person_id_1'], "UNKNOWN")
            },
            "person_2": {
                "id": link['person_id_2'],
                "name": person_map.get(link['person_id_2'], "UNKNOWN")
            },
            "description": f"{person_map.get(link['person_id_1'], '?')} est {link['link_type']} de {person_map.get(link['person_id_2'], '?')}"
        })
    
    return {
        "persons": [{"id": str(p['_id']), "name": f"{p['first_name']} {p['last_name']}"} for p in persons],
        "links": annotated_links,
        "summary": {
            "total_persons": len(persons),
            "total_links": len(links),
            "parent_links": len([l for l in links if l['link_type'] == 'parent']),
            "spouse_links": len([l for l in links if l['link_type'] == 'spouse']),
        }
    }

# ===================== TREE EXPORT ROUTES =====================

@api_router.get("/tree/export/json")
async def export_tree_json(current_user: dict = Depends(get_current_user)):
    """Export family tree as JSON data"""
    user_id = str(current_user['_id'])
    
    # Get all tree data
    persons = await db.persons.find({"user_id": user_id}).to_list(500)
    links = await db.family_links.find({"user_id": user_id}).to_list(500)
    
    tree_data = {
        "app": "AÏLA Family Tree",
        "version": "1.0",
        "exported_at": datetime.utcnow().isoformat(),
        "exported_by": f"{current_user['first_name']} {current_user['last_name']}",
        "persons": [serialize_object_id(p) for p in persons],
        "family_links": [serialize_object_id(l) for l in links],
        "stats": {
            "total_persons": len(persons),
            "total_links": len(links)
        }
    }
    
    return tree_data

@api_router.get("/tree/export/gedcom")
async def export_tree_gedcom(current_user: dict = Depends(get_current_user)):
    """Export family tree as GEDCOM file (genealogy standard format)"""
    from fastapi.responses import Response
    
    user_id = str(current_user['_id'])
    
    # Get all tree data
    persons = await db.persons.find({"user_id": user_id}).to_list(500)
    links = await db.family_links.find({"user_id": user_id}).to_list(500)
    
    # Generate GEDCOM format
    gedcom_lines = [
        "0 HEAD",
        "1 SOUR AÏLA",
        "2 VERS 1.0",
        "2 NAME AÏLA Family Tree",
        "1 DEST ANY",
        "1 DATE " + datetime.utcnow().strftime("%d %b %Y").upper(),
        "1 CHAR UTF-8",
        "1 GEDC",
        "2 VERS 5.5.1",
        "2 FORM LINEAGE-LINKED",
    ]
    
    # Add persons as individuals
    person_id_map = {}
    for idx, person in enumerate(persons, start=1):
        person_id_map[str(person['_id'])] = f"I{idx}"
        
        gedcom_lines.append(f"0 @I{idx}@ INDI")
        gedcom_lines.append(f"1 NAME {person['first_name']} /{person['last_name']}/")
        
        if person.get('gender'):
            gender_code = 'M' if person['gender'] == 'male' else 'F'
            gedcom_lines.append(f"1 SEX {gender_code}")
        
        if person.get('birth_date'):
            gedcom_lines.append(f"1 BIRT")
            gedcom_lines.append(f"2 DATE {person['birth_date']}")
        
        if person.get('birth_place'):
            if not person.get('birth_date'):
                gedcom_lines.append(f"1 BIRT")
            gedcom_lines.append(f"2 PLAC {person['birth_place']}")
        
        if person.get('death_date'):
            gedcom_lines.append(f"1 DEAT")
            gedcom_lines.append(f"2 DATE {person['death_date']}")
    
    # Add family links
    family_idx = 1
    for link in links:
        person1_id = person_id_map.get(str(link['person1_id']))
        person2_id = person_id_map.get(str(link['person2_id']))
        
        if not person1_id or not person2_id:
            continue
        
        if link['link_type'] == 'spouse':
            gedcom_lines.append(f"0 @F{family_idx}@ FAM")
            gedcom_lines.append(f"1 HUSB @{person1_id}@")
            gedcom_lines.append(f"1 WIFE @{person2_id}@")
            family_idx += 1
        elif link['link_type'] in ['parent', 'child']:
            gedcom_lines.append(f"0 @F{family_idx}@ FAM")
            if link['link_type'] == 'parent':
                gedcom_lines.append(f"1 HUSB @{person1_id}@")
                gedcom_lines.append(f"1 CHIL @{person2_id}@")
            else:  # child
                gedcom_lines.append(f"1 CHIL @{person1_id}@")
                gedcom_lines.append(f"1 HUSB @{person2_id}@")
            family_idx += 1
    
    gedcom_lines.append("0 TRLR")
    
    gedcom_content = "\n".join(gedcom_lines)
    
    return Response(
        content=gedcom_content,
        media_type="text/x-gedcom",
        headers={
            "Content-Disposition": f"attachment; filename=aila_tree_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.ged"
        }
    )

# ===================== IMPORT GEDCOM ROUTES =====================

class ImportGedcomRequest(BaseModel):
    gedcom_content: str

@api_router.post("/tree/import/gedcom")
async def import_tree_gedcom(request: ImportGedcomRequest, current_user: dict = Depends(get_current_user)):
    """Import a GEDCOM file and add persons/links to the tree"""
    user_id = str(current_user['_id'])
    
    lines = request.gedcom_content.strip().split('\n')
    
    persons_created = 0
    links_created = 0
    current_person = None
    persons_map = {}  # GEDCOM ID -> MongoDB ID
    families = {}  # Family ID -> {husb, wife, children}
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        parts = line.split(' ', 2)
        
        if len(parts) < 2:
            i += 1
            continue
        
        level = parts[0]
        tag_or_id = parts[1]
        value = parts[2] if len(parts) > 2 else ''
        
        # Parse individual
        if level == '0' and tag_or_id.startswith('@I') and 'INDI' in value:
            gedcom_id = tag_or_id.strip('@')
            current_person = {
                'gedcom_id': gedcom_id,
                'first_name': 'Inconnu',
                'last_name': '',
                'gender': 'unknown',
                'birth_date': None,
                'birth_place': None,
                'death_date': None,
                'death_place': None,
                'user_id': user_id,
            }
        
        elif level == '1' and current_person:
            if tag_or_id == 'NAME' and value:
                # Parse name like "John /Doe/"
                name_parts = value.replace('/', ' ').split()
                if name_parts:
                    current_person['first_name'] = name_parts[0]
                    if len(name_parts) > 1:
                        current_person['last_name'] = ' '.join(name_parts[1:])
            elif tag_or_id == 'SEX':
                current_person['gender'] = 'male' if value == 'M' else 'female' if value == 'F' else 'unknown'
            elif tag_or_id == 'BIRT':
                # Look ahead for date/place
                j = i + 1
                while j < len(lines) and lines[j].strip().startswith('2'):
                    sub_parts = lines[j].strip().split(' ', 2)
                    if len(sub_parts) >= 3:
                        if sub_parts[1] == 'DATE':
                            current_person['birth_date'] = sub_parts[2]
                        elif sub_parts[1] == 'PLAC':
                            current_person['birth_place'] = sub_parts[2]
                    j += 1
            elif tag_or_id == 'DEAT':
                j = i + 1
                while j < len(lines) and lines[j].strip().startswith('2'):
                    sub_parts = lines[j].strip().split(' ', 2)
                    if len(sub_parts) >= 3:
                        if sub_parts[1] == 'DATE':
                            current_person['death_date'] = sub_parts[2]
                        elif sub_parts[1] == 'PLAC':
                            current_person['death_place'] = sub_parts[2]
                    j += 1
        
        # Save person when we hit next record
        if level == '0' and current_person and current_person.get('first_name') != 'Inconnu':
            result = await db.persons.insert_one(current_person)
            persons_map[current_person['gedcom_id']] = str(result.inserted_id)
            persons_created += 1
            current_person = None
        
        # Parse family
        if level == '0' and tag_or_id.startswith('@F') and 'FAM' in value:
            fam_id = tag_or_id.strip('@')
            families[fam_id] = {'husb': None, 'wife': None, 'children': []}
        
        if level == '1' and families:
            current_fam_id = list(families.keys())[-1] if families else None
            if current_fam_id:
                if tag_or_id == 'HUSB' and value:
                    families[current_fam_id]['husb'] = value.strip('@')
                elif tag_or_id == 'WIFE' and value:
                    families[current_fam_id]['wife'] = value.strip('@')
                elif tag_or_id == 'CHIL' and value:
                    families[current_fam_id]['children'].append(value.strip('@'))
        
        i += 1
    
    # Save last person
    if current_person and current_person.get('first_name') != 'Inconnu':
        result = await db.persons.insert_one(current_person)
        persons_map[current_person['gedcom_id']] = str(result.inserted_id)
        persons_created += 1
    
    # Create family links
    for fam_id, fam in families.items():
        husb_id = persons_map.get(fam['husb'])
        wife_id = persons_map.get(fam['wife'])
        
        # Create spouse link
        if husb_id and wife_id:
            await db.family_links.insert_one({
                'user_id': user_id,
                'person_id_1': husb_id,
                'person_id_2': wife_id,
                'link_type': 'spouse',
            })
            links_created += 1
        
        # Create parent links
        for child_gedcom_id in fam['children']:
            child_id = persons_map.get(child_gedcom_id)
            if child_id:
                if husb_id:
                    await db.family_links.insert_one({
                        'user_id': user_id,
                        'person_id_1': husb_id,
                        'person_id_2': child_id,
                        'link_type': 'parent',
                    })
                    links_created += 1
                if wife_id:
                    await db.family_links.insert_one({
                        'user_id': user_id,
                        'person_id_1': wife_id,
                        'person_id_2': child_id,
                        'link_type': 'parent',
                    })
                    links_created += 1
    
    return {
        "message": f"Import réussi ! {persons_created} personnes et {links_created} liens créés.",
        "persons_created": persons_created,
        "links_created": links_created
    }

# ===================== SEND TREE BY EMAIL =====================

class SendTreeEmailRequest(BaseModel):
    recipient_emails: List[str]
    format: str = "pdf"  # "json", "pdf", "csv"
    message: Optional[str] = None

@api_router.post("/tree/send-email")
async def send_tree_by_email(request: SendTreeEmailRequest, current_user: dict = Depends(get_current_user)):
    """Send the family tree data by email to specified recipients"""
    import base64
    import json
    
    user_id = str(current_user['_id'])
    user_name = f"{current_user['first_name']} {current_user['last_name']}"
    
    # Get tree data
    persons = await db.persons.find({"user_id": user_id}).to_list(500)
    links = await db.family_links.find({"user_id": user_id}).to_list(500)
    
    if not persons:
        raise HTTPException(status_code=400, detail="Votre arbre est vide")
    
    # Serialize persons and links for export
    persons_data = [serialize_object_id(p) for p in persons]
    links_data = [serialize_object_id(l) for l in links]
    
    # Create persons lookup for link display
    persons_lookup = {p['id']: p for p in persons_data}
    
    # Generate HTML table for persons
    persons_rows = ""
    for p in persons_data:
        gender_text = "Homme" if p.get('gender') == 'male' else "Femme" if p.get('gender') == 'female' else "Autre"
        persons_rows += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{p.get('first_name', '')} {p.get('last_name', '')}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{gender_text}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{p.get('birth_date', '-') or '-'}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{p.get('death_date', '-') or '-'}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{p.get('birth_place', '-') or '-'}</td>
        </tr>
        """
    
    # Generate HTML table for links
    links_rows = ""
    for link in links_data:
        p1 = persons_lookup.get(link.get('person_id_1'), {})
        p2 = persons_lookup.get(link.get('person_id_2'), {})
        p1_name = f"{p1.get('first_name', '')} {p1.get('last_name', '')}".strip() or "Inconnu"
        p2_name = f"{p2.get('first_name', '')} {p2.get('last_name', '')}".strip() or "Inconnu"
        
        link_type = link.get('link_type', 'autre')
        link_labels = {
            'parent': 'Parent de',
            'spouse': 'Conjoint(e)',
            'sibling': 'Frère/Sœur',
            'child': 'Enfant de',
            'grandparent': 'Grand-parent de',
            'grandchild': 'Petit-enfant de'
        }
        link_label = link_labels.get(link_type, link_type)
        
        links_rows += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{p1_name}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{link_label}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{p2_name}</td>
        </tr>
        """
    
    # Prepare tree summary
    tree_summary = f"{len(persons)} personnes • {len(links)} liens familiaux"
    today = datetime.utcnow().strftime("%d/%m/%Y")
    
    custom_message = request.message or ""
    custom_message_html = ""
    if custom_message:
        custom_message_html = f"""
        <div style="background: #f0f0f0; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #D4AF37;">
            <p style="margin: 0; color: #333;"><strong>Message de {user_name} :</strong></p>
            <p style="margin: 10px 0 0 0; color: #555;">{custom_message}</p>
        </div>
        """
    
    # Prepare attachment based on format
    attachment = None
    attachment_filename = ""
    
    if request.format == "json":
        tree_json = json.dumps({
            "exported_by": user_name,
            "exported_at": datetime.utcnow().isoformat(),
            "persons": persons_data,
            "links": links_data
        }, indent=2, ensure_ascii=False, default=str)
        attachment = base64.b64encode(tree_json.encode('utf-8')).decode('utf-8')
        attachment_filename = f"arbre_{user_name.replace(' ', '_')}_{today.replace('/', '-')}.json"
        
    elif request.format == "csv":
        # Create CSV content
        csv_content = "Prénom,Nom,Genre,Date de naissance,Date de décès,Lieu de naissance,Profession,Notes\n"
        for p in persons_data:
            gender_text = "Homme" if p.get('gender') == 'male' else "Femme" if p.get('gender') == 'female' else "Autre"
            csv_content += f'"{p.get("first_name", "")}","{p.get("last_name", "")}","{gender_text}","{p.get("birth_date", "")}","{p.get("death_date", "")}","{p.get("birth_place", "")}","{p.get("occupation", "")}","{(p.get("notes", "") or "").replace(chr(34), chr(39))}"\n'
        
        csv_content += "\n\nLiens familiaux\nPersonne 1,Relation,Personne 2\n"
        for link in links_data:
            p1 = persons_lookup.get(link.get('person_id_1'), {})
            p2 = persons_lookup.get(link.get('person_id_2'), {})
            p1_name = f"{p1.get('first_name', '')} {p1.get('last_name', '')}".strip() or "Inconnu"
            p2_name = f"{p2.get('first_name', '')} {p2.get('last_name', '')}".strip() or "Inconnu"
            link_type = link.get('link_type', 'autre')
            csv_content += f'"{p1_name}","{link_type}","{p2_name}"\n'
        
        # Add BOM for Excel compatibility
        csv_bytes = ('\ufeff' + csv_content).encode('utf-8')
        attachment = base64.b64encode(csv_bytes).decode('utf-8')
        attachment_filename = f"arbre_{user_name.replace(' ', '_')}_{today.replace('/', '-')}.csv"
    
    emails_sent = 0
    errors = []
    
    for email in request.recipient_emails:
        try:
            # Build HTML email with full tree data
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head><meta charset="UTF-8"></head>
            <body style="font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px;">
                <div style="max-width: 800px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                    <!-- Header -->
                    <div style="background: #0A1628; color: white; padding: 30px; text-align: center;">
                        <h1 style="color: #D4AF37; margin: 0;">🌳 AÏLA</h1>
                        <p style="color: #B8C5D6; margin: 10px 0 0 0;">Arbre Généalogique Familial</p>
                    </div>
                    
                    <!-- Content -->
                    <div style="padding: 30px;">
                        <p style="font-size: 16px;">Bonjour,</p>
                        
                        <p style="font-size: 16px;"><strong style="color: #D4AF37;">{user_name}</strong> vous partage son arbre généalogique complet !</p>
                        
                        <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0; text-align: center;">
                            <p style="margin: 0; font-size: 18px; color: #333;"><strong>{tree_summary}</strong></p>
                            <p style="margin: 5px 0 0 0; color: #666; font-size: 12px;">Exporté le {today}</p>
                        </div>
                        
                        {custom_message_html}
                        
                        <!-- Persons Table -->
                        <h2 style="color: #0A1628; border-bottom: 2px solid #D4AF37; padding-bottom: 10px;">👥 Membres de la famille</h2>
                        <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
                            <tr style="background: #0A1628; color: white;">
                                <th style="padding: 10px; text-align: left;">Nom</th>
                                <th style="padding: 10px; text-align: left;">Genre</th>
                                <th style="padding: 10px; text-align: left;">Naissance</th>
                                <th style="padding: 10px; text-align: left;">Décès</th>
                                <th style="padding: 10px; text-align: left;">Lieu</th>
                            </tr>
                            {persons_rows}
                        </table>
                        
                        <!-- Links Table -->
                        <h2 style="color: #0A1628; border-bottom: 2px solid #D4AF37; padding-bottom: 10px;">🔗 Liens familiaux</h2>
                        <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
                            <tr style="background: #0A1628; color: white;">
                                <th style="padding: 10px; text-align: left;">Personne 1</th>
                                <th style="padding: 10px; text-align: left;">Relation</th>
                                <th style="padding: 10px; text-align: left;">Personne 2</th>
                            </tr>
                            {links_rows if links_rows else '<tr><td colspan="3" style="padding: 10px; text-align: center; color: #999;">Aucun lien enregistré</td></tr>'}
                        </table>
                        
                        <!-- Footer CTA -->
                        <div style="text-align: center; margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 8px;">
                            <p style="margin: 0 0 15px 0;">Créez votre propre arbre généalogique gratuitement :</p>
                            <a href="{FRONTEND_URL}" style="display: inline-block; background: #D4AF37; color: #0A1628; padding: 12px 30px; border-radius: 8px; text-decoration: none; font-weight: bold;">
                                Découvrir AÏLA
                            </a>
                        </div>
                    </div>
                    
                    <!-- Footer -->
                    <div style="background: #0A1628; color: #6B7C93; padding: 20px; text-align: center; font-size: 12px;">
                        <p style="margin: 0;">AÏLA - L'arbre généalogique qui connecte votre famille</p>
                        <p style="margin: 5px 0 0 0;">www.aila.family</p>
                    </div>
                </div>
            </body>
            </html>
            """
            
            params = {
                "from": "AÏLA <noreply@aila.family>",
                "to": [email],
                "subject": f"🌳 Arbre généalogique de {user_name} - {tree_summary}",
                "html": html_content,
            }
            
            # Add attachment if JSON or CSV format
            if attachment and attachment_filename:
                content_type = "application/json" if request.format == "json" else "text/csv"
                params["attachments"] = [{
                    "filename": attachment_filename,
                    "content": attachment,
                    "content_type": content_type
                }]
            
            resend.Emails.send(params)
            emails_sent += 1
            logger.info(f"✓ Tree email sent to {email} (format: {request.format})")
            
        except Exception as e:
            logger.error(f"Error sending email to {email}: {e}")
            errors.append(f"{email}: {str(e)}")
    
    return {
        "message": f"{emails_sent} email(s) envoyé(s) avec succès",
        "emails_sent": emails_sent,
        "errors": errors if errors else None
    }

# ===================== TREE MERGE ROUTES =====================

def normalize_name(name: str) -> str:
    """Normalize a name for comparison"""
    if not name:
        return ""
    import unicodedata
    # Remove accents
    normalized = unicodedata.normalize('NFD', name.lower())
    normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
    # Remove extra spaces and special characters
    normalized = ' '.join(normalized.split())
    return normalized.strip()

def parse_date_for_comparison(date_str: Optional[str]) -> Optional[datetime]:
    """Parse a date string for comparison"""
    if not date_str:
        return None
    
    date_str = str(date_str).split("T")[0].strip()
    
    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def calculate_similarity(person1: dict, person2: dict) -> tuple[float, str]:
    """
    Calculate similarity between two persons.
    Returns (score 0-100, reason)
    """
    score = 0
    reasons = []
    
    # Compare first name (40 points max)
    name1 = normalize_name(person1.get('first_name', ''))
    name2 = normalize_name(person2.get('first_name', ''))
    if name1 and name2:
        if name1 == name2:
            score += 40
            reasons.append("Même prénom")
        elif name1 in name2 or name2 in name1:
            score += 25
            reasons.append("Prénom similaire")
    
    # Compare last name (40 points max)
    lastname1 = normalize_name(person1.get('last_name', ''))
    lastname2 = normalize_name(person2.get('last_name', ''))
    if lastname1 and lastname2:
        if lastname1 == lastname2:
            score += 40
            reasons.append("Même nom de famille")
        elif lastname1 in lastname2 or lastname2 in lastname1:
            score += 25
            reasons.append("Nom similaire")
    
    # Compare birth date (20 points max)
    date1 = parse_date_for_comparison(person1.get('birth_date'))
    date2 = parse_date_for_comparison(person2.get('birth_date'))
    if date1 and date2:
        diff_days = abs((date1 - date2).days)
        if diff_days == 0:
            score += 20
            reasons.append("Même date de naissance")
        elif diff_days <= 365:  # Within a year (could be partial dates)
            score += 10
            reasons.append("Année de naissance proche")
    
    # Bonus for same gender
    if person1.get('gender') and person2.get('gender'):
        if person1['gender'] == person2['gender']:
            score = min(100, score + 5)
    
    reason = " + ".join(reasons) if reasons else "Aucune correspondance"
    return (min(100, score), reason)

@api_router.get("/tree/merge/shared-trees")
async def get_mergeable_trees(current_user: dict = Depends(get_current_user)):
    """Get list of trees that can be merged (trees shared with current user as editor)"""
    user_id = str(current_user['_id'])
    
    # Get all collaborations where user is editor
    collaborations = await db.collaborators.find({
        "user_id": user_id,
        "status": "accepted",
        "role": "editor"
    }).to_list(100)
    
    mergeable_trees = []
    for collab in collaborations:
        owner_id = collab['tree_owner_id']
        owner = await db.users.find_one({"_id": ObjectId(owner_id)})
        if owner:
            persons_count = await db.persons.count_documents({"user_id": owner_id})
            mergeable_trees.append({
                "owner_id": owner_id,
                "owner_name": f"{owner['first_name']} {owner['last_name']}",
                "owner_email": owner['email'],
                "persons_count": persons_count,
                "role": collab['role']
            })
    
    return {"trees": mergeable_trees}

@api_router.post("/tree/merge/analyze", response_model=MergeAnalysisResponse)
async def analyze_merge(source_tree_owner_id: str, current_user: dict = Depends(get_current_user)):
    """
    Analyze a potential merge between source tree and current user's tree.
    Detects duplicates and prepares merge plan.
    """
    user_id = str(current_user['_id'])
    
    # Verify user has access to source tree (must be collaborator with editor role)
    collaboration = await db.collaborators.find_one({
        "tree_owner_id": source_tree_owner_id,
        "user_id": user_id,
        "status": "accepted",
        "role": "editor"
    })
    
    if not collaboration:
        raise HTTPException(
            status_code=403, 
            detail="Vous devez être éditeur de l'arbre source pour pouvoir le fusionner"
        )
    
    # Get source tree owner info
    source_owner = await db.users.find_one({"_id": ObjectId(source_tree_owner_id)})
    if not source_owner:
        raise HTTPException(status_code=404, detail="Propriétaire de l'arbre source non trouvé")
    
    # Get source tree persons
    source_persons = await db.persons.find({"user_id": source_tree_owner_id}).to_list(1000)
    
    # Get target tree persons (current user's tree)
    target_persons = await db.persons.find({"user_id": user_id}).to_list(1000)
    
    # Get source tree links
    source_links = await db.family_links.find({"user_id": source_tree_owner_id}).to_list(1000)
    
    # Find duplicates
    duplicates = []
    matched_source_ids = set()
    
    for source_person in source_persons:
        best_match = None
        best_score = 0
        best_reason = ""
        
        for target_person in target_persons:
            score, reason = calculate_similarity(source_person, target_person)
            if score >= 60 and score > best_score:  # Threshold of 60%
                best_score = score
                best_match = target_person
                best_reason = reason
        
        if best_match:
            matched_source_ids.add(str(source_person['_id']))
            duplicates.append(DuplicateCandidate(
                source_person_id=str(source_person['_id']),
                source_person_name=f"{source_person['first_name']} {source_person['last_name']}",
                source_birth_date=source_person.get('birth_date'),
                target_person_id=str(best_match['_id']),
                target_person_name=f"{best_match['first_name']} {best_match['last_name']}",
                target_birth_date=best_match.get('birth_date'),
                similarity_score=best_score,
                match_reason=best_reason
            ))
    
    # Count new persons (not duplicates)
    new_persons_count = len(source_persons) - len(matched_source_ids)
    
    logger.info(f"Merge analysis: {len(source_persons)} source persons, {len(duplicates)} duplicates found, {new_persons_count} new persons")
    
    return MergeAnalysisResponse(
        source_tree_owner_id=source_tree_owner_id,
        source_tree_owner_name=f"{source_owner['first_name']} {source_owner['last_name']}",
        source_persons_count=len(source_persons),
        target_persons_count=len(target_persons),
        duplicates_found=duplicates,
        new_persons_count=new_persons_count,
        new_links_count=len(source_links)
    )

@api_router.post("/tree/merge/execute", response_model=MergeExecuteResponse)
async def execute_merge(request: MergeExecuteRequest, current_user: dict = Depends(get_current_user)):
    """
    Execute the merge based on user decisions.
    - merge: Use existing target person (link source data to target)
    - add: Add source person as new person in target tree
    - skip: Don't import this person
    """
    user_id = str(current_user['_id'])
    
    # Verify user has access to source tree
    collaboration = await db.collaborators.find_one({
        "tree_owner_id": request.source_tree_owner_id,
        "user_id": user_id,
        "status": "accepted",
        "role": "editor"
    })
    
    if not collaboration:
        raise HTTPException(
            status_code=403, 
            detail="Vous devez être éditeur de l'arbre source pour pouvoir le fusionner"
        )
    
    # Get source tree data
    source_persons = await db.persons.find({"user_id": request.source_tree_owner_id}).to_list(1000)
    source_links = await db.family_links.find({"user_id": request.source_tree_owner_id}).to_list(1000)
    
    # Build source person map
    source_person_map = {str(p['_id']): p for p in source_persons}
    
    # Build ID mapping: source_id -> target_id
    id_mapping = {}
    
    persons_merged = 0
    persons_added = 0
    persons_skipped = 0
    
    # Process decisions
    decision_map = {d.source_person_id: d for d in request.decisions}
    
    for source_person in source_persons:
        source_id = str(source_person['_id'])
        decision = decision_map.get(source_id)
        
        if decision:
            if decision.action == "merge" and decision.target_person_id:
                # Map source ID to existing target ID
                id_mapping[source_id] = decision.target_person_id
                persons_merged += 1
                logger.info(f"Merging {source_person['first_name']} -> {decision.target_person_id}")
                
            elif decision.action == "add":
                # Create new person in target tree
                new_person = {
                    "user_id": user_id,
                    "first_name": source_person['first_name'],
                    "last_name": source_person['last_name'],
                    "gender": source_person.get('gender', 'unknown'),
                    "birth_date": source_person.get('birth_date'),
                    "birth_place": source_person.get('birth_place'),
                    "death_date": source_person.get('death_date'),
                    "death_place": source_person.get('death_place'),
                    "photo": source_person.get('photo'),
                    "notes": source_person.get('notes'),
                    "geographic_branch": source_person.get('geographic_branch'),
                    "created_at": datetime.utcnow(),
                    "is_preview": False,
                    "merged_from": request.source_tree_owner_id  # Track origin
                }
                result = await db.persons.insert_one(new_person)
                id_mapping[source_id] = str(result.inserted_id)
                persons_added += 1
                logger.info(f"Added new person: {source_person['first_name']} {source_person['last_name']}")
                
            elif decision.action == "skip":
                persons_skipped += 1
                logger.info(f"Skipped: {source_person['first_name']} {source_person['last_name']}")
        else:
            # No decision provided - add as new by default
            new_person = {
                "user_id": user_id,
                "first_name": source_person['first_name'],
                "last_name": source_person['last_name'],
                "gender": source_person.get('gender', 'unknown'),
                "birth_date": source_person.get('birth_date'),
                "birth_place": source_person.get('birth_place'),
                "death_date": source_person.get('death_date'),
                "death_place": source_person.get('death_place'),
                "photo": source_person.get('photo'),
                "notes": source_person.get('notes'),
                "geographic_branch": source_person.get('geographic_branch'),
                "created_at": datetime.utcnow(),
                "is_preview": False,
                "merged_from": request.source_tree_owner_id
            }
            result = await db.persons.insert_one(new_person)
            id_mapping[source_id] = str(result.inserted_id)
            persons_added += 1
    
    # Import links if requested
    links_added = 0
    if request.import_links:
        for source_link in source_links:
            source_person1_id = source_link['person_id_1']
            source_person2_id = source_link['person_id_2']
            
            # Map to target IDs
            target_person1_id = id_mapping.get(source_person1_id)
            target_person2_id = id_mapping.get(source_person2_id)
            
            # Only create link if both persons exist in target tree
            if target_person1_id and target_person2_id:
                # Check if link already exists
                existing_link = await db.family_links.find_one({
                    "user_id": user_id,
                    "$or": [
                        {"person_id_1": target_person1_id, "person_id_2": target_person2_id, "link_type": source_link['link_type']},
                        {"person_id_1": target_person2_id, "person_id_2": target_person1_id, "link_type": source_link['link_type']}
                    ]
                })
                
                if not existing_link:
                    new_link = {
                        "user_id": user_id,
                        "person_id_1": target_person1_id,
                        "person_id_2": target_person2_id,
                        "link_type": source_link['link_type'],
                        "created_at": datetime.utcnow()
                    }
                    await db.family_links.insert_one(new_link)
                    links_added += 1
    
    message = f"Fusion terminée ! {persons_merged} personne(s) fusionnée(s), {persons_added} personne(s) ajoutée(s), {links_added} lien(s) créé(s)"
    if persons_skipped > 0:
        message += f", {persons_skipped} personne(s) ignorée(s)"
    
    logger.info(message)
    
    return MergeExecuteResponse(
        persons_merged=persons_merged,
        persons_added=persons_added,
        persons_skipped=persons_skipped,
        links_added=links_added,
        message=message
    )

# ===================== COLLABORATION ROUTES =====================

@api_router.post("/collaborators/invite", response_model=CollaboratorResponse)
async def invite_collaborator(invite: CollaboratorInvite, current_user: dict = Depends(get_current_user)):
    """Invite someone to collaborate on your family tree"""
    user_id = str(current_user['_id'])
    
    # Normalize email to lowercase for consistent comparison
    invite_email = invite.email.lower().strip()
    user_email = current_user['email'].lower().strip()
    
    # Check if already invited (case-insensitive)
    existing = await db.collaborators.find_one({
        "tree_owner_id": user_id,
        "email": {"$regex": f"^{invite_email}$", "$options": "i"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Cette personne a déjà été invitée")
    
    # Check if inviting yourself
    if invite_email == user_email:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas vous inviter vous-même")
    
    # Check if the email is registered (case-insensitive)
    invited_user = await db.users.find_one({"email": {"$regex": f"^{invite_email}$", "$options": "i"}})
    
    # Create invitation token
    invite_token = str(uuid.uuid4())
    
    collaborator_doc = {
        "tree_owner_id": user_id,
        "user_id": str(invited_user['_id']) if invited_user else None,
        "email": invite_email,  # Store normalized email
        "role": invite.role,
        "status": "accepted" if invited_user else "pending",  # Auto-accept if user exists
        "invite_token": invite_token,
        "invited_at": datetime.utcnow(),
        "accepted_at": datetime.utcnow() if invited_user else None
    }
    
    result = await db.collaborators.insert_one(collaborator_doc)
    collaborator_doc['_id'] = result.inserted_id
    collaborator_doc = serialize_object_id(collaborator_doc)
    
    # Send invitation email (use original email for display)
    owner_name = f"{current_user['first_name']} {current_user['last_name']}"
    await send_invitation_email(
        to_email=invite.email,  # Use original email for sending
        inviter_name=owner_name,
        role=invite.role,
        invite_token=invite_token
    )
    
    # Create notification for the invited user if they exist
    if invited_user:
        await db.notifications.insert_one({
            "user_id": str(invited_user['_id']),
            "type": "collaboration_invite",
            "title": "Nouvelle invitation",
            "message": f"{owner_name} vous a invité à collaborer sur son arbre généalogique",
            "data": {"tree_owner_id": user_id, "collaborator_id": collaborator_doc['id']},
            "read": False,
            "created_at": datetime.utcnow()
        })
    
    return CollaboratorResponse(
        id=collaborator_doc['id'],
        tree_owner_id=collaborator_doc['tree_owner_id'],
        user_id=collaborator_doc.get('user_id'),
        email=collaborator_doc['email'],
        role=collaborator_doc['role'],
        status=collaborator_doc['status'],
        invited_at=collaborator_doc['invited_at'],
        accepted_at=collaborator_doc.get('accepted_at')
    )

@api_router.get("/collaborators", response_model=List[CollaboratorResponse])
async def get_collaborators(current_user: dict = Depends(get_current_user)):
    """Get all collaborators for your tree"""
    user_id = str(current_user['_id'])
    collaborators = await db.collaborators.find({"tree_owner_id": user_id}).to_list(100)
    return [CollaboratorResponse(
        id=str(c['_id']),
        tree_owner_id=c['tree_owner_id'],
        user_id=c.get('user_id'),
        email=c['email'],
        role=c['role'],
        status=c['status'],
        invited_at=c['invited_at'],
        accepted_at=c.get('accepted_at')
    ) for c in collaborators]

@api_router.get("/collaborators/shared-with-me")
async def get_trees_shared_with_me(current_user: dict = Depends(get_current_user)):
    """Get trees that others have shared with you"""
    user_id = str(current_user['_id'])
    collaborations = await db.collaborators.find({
        "user_id": user_id,
        "status": "accepted"
    }).to_list(100)
    
    result = []
    for collab in collaborations:
        owner = await db.users.find_one({"_id": ObjectId(collab['tree_owner_id'])})
        if owner:
            persons_count = await db.persons.count_documents({"user_id": collab['tree_owner_id']})
            result.append({
                "id": str(collab['_id']),
                "owner_id": collab['tree_owner_id'],
                "owner_name": f"{owner['first_name']} {owner['last_name']}",
                "owner_email": owner['email'],
                "role": collab['role'],
                "persons_count": persons_count,
                "accepted_at": collab.get('accepted_at')
            })
    
    return result

@api_router.delete("/collaborators/{collaborator_id}")
async def remove_collaborator(collaborator_id: str, current_user: dict = Depends(get_current_user)):
    """Remove a collaborator from your tree"""
    user_id = str(current_user['_id'])
    
    result = await db.collaborators.delete_one({
        "_id": ObjectId(collaborator_id),
        "tree_owner_id": user_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Collaborator not found")
    
    return {"message": "Collaborator removed successfully"}

@api_router.post("/collaborators/accept/{invite_token}")
async def accept_invitation(invite_token: str, current_user: dict = Depends(get_current_user)):
    """Accept a collaboration invitation using the token"""
    user_email = current_user['email'].lower().strip()  # Normalize email
    user_id = str(current_user['_id'])
    
    # Find the invitation by token
    invitation = await db.collaborators.find_one({"invite_token": invite_token})
    
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found")
    
    # Compare emails case-insensitively
    invitation_email = invitation['email'].lower().strip()
    if invitation_email != user_email:
        raise HTTPException(status_code=403, detail=f"Cette invitation a été envoyée à une autre adresse email ({invitation['email']})")
    
    if invitation['status'] == 'accepted':
        # Already accepted, return the tree owner info
        owner = await db.users.find_one({"_id": ObjectId(invitation['tree_owner_id'])})
        return {
            "message": "Invitation already accepted",
            "tree_owner_id": invitation['tree_owner_id'],
            "tree_owner_name": f"{owner['first_name']} {owner['last_name']}" if owner else "Unknown",
            "role": invitation['role']
        }
    
    # Accept the invitation
    await db.collaborators.update_one(
        {"_id": invitation['_id']},
        {
            "$set": {
                "status": "accepted",
                "user_id": user_id,
                "accepted_at": datetime.utcnow()
            }
        }
    )
    
    # Get tree owner info
    owner = await db.users.find_one({"_id": ObjectId(invitation['tree_owner_id'])})
    
    return {
        "message": "Invitation accepted successfully",
        "tree_owner_id": invitation['tree_owner_id'],
        "tree_owner_name": f"{owner['first_name']} {owner['last_name']}" if owner else "Unknown",
        "role": invitation['role']
    }

@api_router.get("/tree/shared/{owner_id}", response_model=TreeResponse)
async def get_shared_tree(owner_id: str, current_user: dict = Depends(get_current_user)):
    """Get a shared tree that you have access to"""
    user_id = str(current_user['_id'])
    
    # Check if user has access
    collaboration = await db.collaborators.find_one({
        "tree_owner_id": owner_id,
        "user_id": user_id,
        "status": "accepted"
    })
    
    if not collaboration:
        raise HTTPException(status_code=403, detail="You don't have access to this tree")
    
    persons = await db.persons.find({"user_id": owner_id}).to_list(500)
    links = await db.family_links.find({"user_id": owner_id}).to_list(500)
    
    return TreeResponse(
        persons=[PersonResponse(**serialize_object_id(p)) for p in persons],
        links=[FamilyLinkResponse(**serialize_object_id(l)) for l in links]
    )

# ===================== SHARED TREE EDITOR ROUTES =====================

@api_router.post("/tree/shared/{owner_id}/persons", response_model=PersonResponse)
async def create_person_in_shared_tree(owner_id: str, person: PersonCreate, current_user: dict = Depends(get_current_user)):
    """Create a person in a shared tree (editor only)"""
    user_id = str(current_user['_id'])
    
    # Check if user has editor access
    collaboration = await db.collaborators.find_one({
        "tree_owner_id": owner_id,
        "user_id": user_id,
        "status": "accepted",
        "role": "editor"
    })
    
    if not collaboration:
        raise HTTPException(status_code=403, detail="Vous n'avez pas les droits d'éditeur sur cet arbre")
    
    # Create person in the owner's tree
    person_doc = {
        "user_id": owner_id,  # Important: use owner_id, not current user
        "first_name": person.first_name,
        "last_name": person.last_name,
        "gender": person.gender,
        "birth_date": person.birth_date,
        "birth_place": person.birth_place,
        "death_date": person.death_date,
        "death_place": person.death_place,
        "geographic_branch": person.geographic_branch,
        "notes": person.notes,
        "created_by": user_id,  # Track who created it
        "created_at": datetime.now().isoformat()
    }
    
    result = await db.persons.insert_one(person_doc)
    person_doc['_id'] = result.inserted_id
    person_doc = serialize_object_id(person_doc)
    
    return PersonResponse(**person_doc)

@api_router.post("/tree/shared/{owner_id}/links", response_model=FamilyLinkResponse)
async def create_link_in_shared_tree(owner_id: str, link: FamilyLinkCreate, current_user: dict = Depends(get_current_user)):
    """Create a family link in a shared tree (editor only)"""
    user_id = str(current_user['_id'])
    
    # Check if user has editor access
    collaboration = await db.collaborators.find_one({
        "tree_owner_id": owner_id,
        "user_id": user_id,
        "status": "accepted",
        "role": "editor"
    })
    
    if not collaboration:
        raise HTTPException(status_code=403, detail="Vous n'avez pas les droits d'éditeur sur cet arbre")
    
    # Verify both persons exist in the owner's tree
    person1 = await db.persons.find_one({"_id": ObjectId(link.person_id_1), "user_id": owner_id})
    person2 = await db.persons.find_one({"_id": ObjectId(link.person_id_2), "user_id": owner_id})
    
    if not person1 or not person2:
        raise HTTPException(status_code=404, detail="Une ou plusieurs personnes n'existent pas dans cet arbre")
    
    # Check for existing link
    existing = await db.family_links.find_one({
        "user_id": owner_id,
        "$or": [
            {"person_id_1": link.person_id_1, "person_id_2": link.person_id_2, "link_type": link.link_type},
            {"person_id_1": link.person_id_2, "person_id_2": link.person_id_1, "link_type": link.link_type}
        ]
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Ce lien existe déjà")
    
    # Create link in the owner's tree
    link_doc = {
        "user_id": owner_id,  # Important: use owner_id
        "person_id_1": link.person_id_1,
        "person_id_2": link.person_id_2,
        "link_type": link.link_type,
        "created_by": user_id,  # Track who created it
        "created_at": datetime.now().isoformat()
    }
    
    result = await db.family_links.insert_one(link_doc)
    link_doc['_id'] = result.inserted_id
    link_doc = serialize_object_id(link_doc)
    
    return FamilyLinkResponse(**link_doc)

# ===================== CONTRIBUTIONS ROUTES =====================

@api_router.post("/contributions", response_model=ContributionResponse)
async def create_contribution(contribution: ContributionCreate, owner_id: str, current_user: dict = Depends(get_current_user)):
    """Create a contribution (modification request) for a shared tree"""
    user_id = str(current_user['_id'])
    
    # Check if user has editor access
    collaboration = await db.collaborators.find_one({
        "tree_owner_id": owner_id,
        "user_id": user_id,
        "status": "accepted",
        "role": "editor"
    })
    
    if not collaboration:
        raise HTTPException(status_code=403, detail="You don't have editor access to this tree")
    
    contributor_name = f"{current_user['first_name']} {current_user['last_name']}"
    
    contribution_doc = {
        "tree_owner_id": owner_id,
        "contributor_id": user_id,
        "contributor_name": contributor_name,
        "action": contribution.action,
        "entity_type": contribution.entity_type,
        "entity_id": contribution.entity_id,
        "entity_data": contribution.entity_data,
        "status": "pending",
        "created_at": datetime.utcnow(),
        "reviewed_at": None,
        "review_note": None
    }
    
    result = await db.contributions.insert_one(contribution_doc)
    contribution_doc['_id'] = result.inserted_id
    contribution_doc = serialize_object_id(contribution_doc)
    
    # Notify the tree owner
    await db.notifications.insert_one({
        "user_id": owner_id,
        "type": "new_contribution",
        "title": "Nouvelle contribution",
        "message": f"{contributor_name} a proposé une modification à votre arbre",
        "data": {"contribution_id": contribution_doc['id']},
        "read": False,
        "created_at": datetime.utcnow()
    })
    
    return ContributionResponse(**contribution_doc)

@api_router.get("/contributions/pending", response_model=List[ContributionResponse])
async def get_pending_contributions(current_user: dict = Depends(get_current_user)):
    """Get all pending contributions for your tree"""
    user_id = str(current_user['_id'])
    
    contributions = await db.contributions.find({
        "tree_owner_id": user_id,
        "status": "pending"
    }).to_list(100)
    
    return [ContributionResponse(**serialize_object_id(c)) for c in contributions]

@api_router.get("/contributions/my", response_model=List[ContributionResponse])
async def get_my_contributions(current_user: dict = Depends(get_current_user)):
    """Get all contributions you've made to other trees"""
    user_id = str(current_user['_id'])
    
    contributions = await db.contributions.find({
        "contributor_id": user_id
    }).sort("created_at", -1).to_list(100)
    
    return [ContributionResponse(**serialize_object_id(c)) for c in contributions]

@api_router.post("/contributions/{contribution_id}/review")
async def review_contribution(contribution_id: str, review: ContributionReview, current_user: dict = Depends(get_current_user)):
    """Approve or reject a contribution"""
    user_id = str(current_user['_id'])
    
    contribution = await db.contributions.find_one({
        "_id": ObjectId(contribution_id),
        "tree_owner_id": user_id,
        "status": "pending"
    })
    
    if not contribution:
        raise HTTPException(status_code=404, detail="Contribution not found or already reviewed")
    
    # If approved, apply the change
    if review.status == "approved":
        if contribution['action'] == "add" and contribution['entity_type'] == "person":
            person_data = contribution['entity_data']
            person_data['user_id'] = user_id
            person_data['created_at'] = datetime.utcnow()
            person_data['is_preview'] = False
            await db.persons.insert_one(person_data)
        
        elif contribution['action'] == "add" and contribution['entity_type'] == "link":
            link_data = contribution['entity_data']
            link_data['user_id'] = user_id
            link_data['created_at'] = datetime.utcnow()
            await db.family_links.insert_one(link_data)
        
        elif contribution['action'] == "edit" and contribution['entity_type'] == "person":
            await db.persons.update_one(
                {"_id": ObjectId(contribution['entity_id']), "user_id": user_id},
                {"$set": contribution['entity_data']}
            )
        
        elif contribution['action'] == "delete" and contribution['entity_type'] == "person":
            await db.persons.delete_one({"_id": ObjectId(contribution['entity_id']), "user_id": user_id})
            await db.family_links.delete_many({
                "$or": [
                    {"person_id_1": contribution['entity_id']},
                    {"person_id_2": contribution['entity_id']}
                ],
                "user_id": user_id
            })
    
    # Update contribution status
    await db.contributions.update_one(
        {"_id": ObjectId(contribution_id)},
        {"$set": {
            "status": review.status,
            "reviewed_at": datetime.utcnow(),
            "review_note": review.note
        }}
    )
    
    # Notify contributor
    status_text = "approuvée" if review.status == "approved" else "refusée"
    await db.notifications.insert_one({
        "user_id": contribution['contributor_id'],
        "type": "contribution_reviewed",
        "title": f"Contribution {status_text}",
        "message": f"Votre contribution a été {status_text}" + (f": {review.note}" if review.note else ""),
        "data": {"contribution_id": contribution_id, "status": review.status},
        "read": False,
        "created_at": datetime.utcnow()
    })
    
    return {"message": f"Contribution {review.status}", "contribution_id": contribution_id}

# ===================== NOTIFICATIONS ROUTES =====================

@api_router.get("/notifications", response_model=List[NotificationResponse])
async def get_notifications(current_user: dict = Depends(get_current_user)):
    """Get all notifications for the current user"""
    user_id = str(current_user['_id'])
    
    notifications = await db.notifications.find({
        "user_id": user_id
    }).sort("created_at", -1).to_list(50)
    
    return [NotificationResponse(**serialize_object_id(n)) for n in notifications]

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: dict = Depends(get_current_user)):
    """Get count of unread notifications"""
    user_id = str(current_user['_id'])
    count = await db.notifications.count_documents({"user_id": user_id, "read": False})
    return {"unread_count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark a notification as read"""
    user_id = str(current_user['_id'])
    
    result = await db.notifications.update_one(
        {"_id": ObjectId(notification_id), "user_id": user_id},
        {"$set": {"read": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Mark all notifications as read"""
    user_id = str(current_user['_id'])
    
    await db.notifications.update_many(
        {"user_id": user_id, "read": False},
        {"$set": {"read": True}}
    )
    
    return {"message": "All notifications marked as read"}

# ===================== CHAT ROUTES =====================
# IMPORTANT: Chat messages are STRICTLY PRIVATE per user's own tree
# - Each user sees ONLY their own tree's messages on their session
# - To see a shared tree's chat, user must explicitly view that tree (via tree_id parameter)

@api_router.post("/chat/messages", response_model=ChatMessageResponse)
async def send_message(
    message_data: ChatMessageCreate,
    tree_id: Optional[str] = None,  # Optional: specify which tree to post to
    current_user: dict = Depends(get_current_user)
):
    """
    Send a chat message.
    - If tree_id is not provided: posts to YOUR OWN tree
    - If tree_id is provided: posts to that tree (must have editor access)
    """
    user_id = str(current_user['_id'])
    user_name = f"{current_user['first_name']} {current_user['last_name']}"
    
    # Determine which tree this message belongs to
    if tree_id and tree_id != user_id:
        # Posting to a shared tree - verify editor access
        collaboration = await db.collaborators.find_one({
            "tree_owner_id": tree_id,
            "user_id": user_id,
            "status": "accepted",
            "role": "editor"
        })
        if not collaboration:
            raise HTTPException(status_code=403, detail="Vous n'avez pas les droits pour poster dans ce chat")
        target_tree_id = tree_id
    else:
        # Posting to own tree
        target_tree_id = user_id
    
    # Prepare message document
    message_doc = {
        "tree_owner_id": target_tree_id,  # CRITICAL: Associates message with a specific tree
        "user_id": user_id,
        "user_name": user_name,
        "message": message_data.message,
        "mentioned_person_id": message_data.mentioned_person_id,
        "mentioned_person_name": None,
        "created_at": datetime.utcnow()
    }
    
    # If a person is mentioned, get their name
    if message_data.mentioned_person_id:
        try:
            person = await db.persons.find_one({"_id": ObjectId(message_data.mentioned_person_id)})
            if person:
                message_doc["mentioned_person_name"] = f"{person['first_name']} {person['last_name']}"
        except:
            pass
    
    # Insert message
    result = await db.chat_messages.insert_one(message_doc)
    message_doc['_id'] = result.inserted_id
    
    logger.info(f"Chat message sent by {user_name} to tree: {target_tree_id}")
    
    return ChatMessageResponse(**serialize_object_id(message_doc))

@api_router.get("/chat/messages", response_model=List[ChatMessageResponse])
async def get_messages(
    limit: int = 50,
    skip: int = 0,
    tree_id: Optional[str] = None,  # Optional: specify which tree's chat to view
    current_user: dict = Depends(get_current_user)
):
    """
    Get chat messages - STRICTLY PRIVATE per tree.
    - If tree_id is not provided: returns ONLY YOUR OWN tree's messages
    - If tree_id is provided: returns that tree's messages (must have access)
    """
    user_id = str(current_user['_id'])
    
    # Determine which tree's messages to fetch
    if tree_id and tree_id != user_id:
        # Viewing a shared tree's chat - verify access
        collaboration = await db.collaborators.find_one({
            "tree_owner_id": tree_id,
            "user_id": user_id,
            "status": "accepted"
        })
        if not collaboration:
            raise HTTPException(status_code=403, detail="Vous n'avez pas accès à ce chat")
        target_tree_id = tree_id
        logger.info(f"User {user_id} viewing shared tree chat: {tree_id}")
    else:
        # Viewing own tree's chat
        target_tree_id = user_id
        logger.info(f"User {user_id} viewing own tree chat")
    
    # STRICT PRIVACY: Only fetch messages from the specified tree
    query = {
        "$or": [
            {"tree_owner_id": target_tree_id},
            # Legacy messages (no tree_owner_id) - only show if viewing own tree
            {"tree_owner_id": {"$exists": False}, "user_id": target_tree_id}
        ]
    }
    
    messages = await db.chat_messages.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Reverse to show oldest first
    messages.reverse()
    
    logger.info(f"Returning {len(messages)} messages for tree: {target_tree_id}")
    
    return [ChatMessageResponse(**serialize_object_id(m)) for m in messages]

@api_router.delete("/chat/messages/{message_id}")
async def delete_message(
    message_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a chat message (only your own messages or if admin)"""
    user_id = str(current_user['_id'])
    is_admin = current_user.get('role') == 'admin'
    
    message = await db.chat_messages.find_one({"_id": ObjectId(message_id)})
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    # Check permissions - must be message author or admin
    if not is_admin and message['user_id'] != user_id:
        raise HTTPException(status_code=403, detail="You can only delete your own messages")
    
    await db.chat_messages.delete_one({"_id": ObjectId(message_id)})
    
    return {"message": "Message deleted successfully"}

@api_router.post("/chat/migrate-messages")
async def migrate_chat_messages(current_user: dict = Depends(get_current_user)):
    """
    Migration endpoint: Fix all chat messages to have correct tree_owner_id.
    Only admin can run this.
    """
    if current_user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Find all messages without tree_owner_id or with wrong tree_owner_id
    messages_without_tree = await db.chat_messages.find({
        "$or": [
            {"tree_owner_id": {"$exists": False}},
            {"tree_owner_id": None}
        ]
    }).to_list(1000)
    
    fixed_count = 0
    for msg in messages_without_tree:
        # Set tree_owner_id to the message author's user_id
        await db.chat_messages.update_one(
            {"_id": msg['_id']},
            {"$set": {"tree_owner_id": msg['user_id']}}
        )
        fixed_count += 1
    
    return {
        "message": f"Migration complete. Fixed {fixed_count} messages.",
        "fixed_count": fixed_count
    }

# ===================== FAMILY EVENTS =====================

@api_router.get("/events/birthdays", response_model=List[UpcomingBirthdayResponse])
async def get_upcoming_birthdays(current_user: dict = Depends(get_current_user)):
    """Get upcoming birthdays from tree members (next 7 days) - EXCLUDING DECEASED PERSONS"""
    user_id = str(current_user['_id'])
    
    # Get all persons with birth dates, EXCLUDING deceased persons
    # A person is deceased if death_date exists and is not empty
    persons = await db.persons.find({
        "user_id": user_id,
        "birth_date": {"$ne": None, "$exists": True},
        "$or": [
            {"death_date": {"$exists": False}},  # No death_date field
            {"death_date": None},                 # death_date is null
            {"death_date": ""}                    # death_date is empty string
        ]
    }).to_list(None)
    
    upcoming = []
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    def parse_birth_date(date_str: str):
        """Parse birth date in multiple formats"""
        if not date_str:
            return None
        
        # Remove time component if present
        date_str = date_str.split("T")[0].strip()
        
        # Try different formats
        formats = [
            "%Y-%m-%d",      # 1964-12-25
            "%d/%m/%Y",      # 25/12/1964
            "%d-%m-%Y",      # 25-12-1964
            "%Y",            # 1964 (year only)
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        return None
    
    for person in persons:
        birth_date_str = person.get("birth_date")
        if not birth_date_str:
            continue
        
        # Double-check: Skip if person has ANY death date (even partial)
        death_date = person.get("death_date")
        if death_date and str(death_date).strip():
            logger.info(f"Skipping birthday for deceased person: {person.get('first_name')} {person.get('last_name')}")
            continue
        
        birth_date = parse_birth_date(birth_date_str)
        if not birth_date:
            logger.warning(f"Could not parse birth date: {birth_date_str}")
            continue
        
        try:
            # Calculate this year's birthday
            this_year_birthday = birth_date.replace(year=today.year)
            if this_year_birthday < today:
                this_year_birthday = this_year_birthday.replace(year=today.year + 1)
            
            days_until = (this_year_birthday - today).days
            
            # Only include birthdays in the next 7 days (popup reminder)
            if 0 <= days_until <= 7:
                age = this_year_birthday.year - birth_date.year
                
                upcoming.append(UpcomingBirthdayResponse(
                    person_id=str(person["_id"]),
                    person_name=f"{person['first_name']} {person['last_name']}",
                    birth_date=birth_date_str,
                    days_until=days_until,
                    age=age
                ))
        except (ValueError, TypeError) as e:
            logger.error(f"Error processing birthday for {person.get('first_name', 'Unknown')}: {e}")
            continue
    
    # Sort by days until birthday
    upcoming.sort(key=lambda x: x.days_until)
    logger.info(f"Found {len(upcoming)} upcoming birthdays in next 7 days (excluding deceased)")
    return upcoming

@api_router.get("/events/today")
async def get_todays_events(current_user: dict = Depends(get_current_user)):
    """Get today's birthdays and events for animation display - EXCLUDING DECEASED PERSONS"""
    user_id = str(current_user['_id'])
    today = datetime.now()
    today_str = today.strftime("%m-%d")
    
    events = []
    
    # Check for birthdays today - EXCLUDING deceased persons
    persons = await db.persons.find({
        "user_id": user_id,
        "birth_date": {"$ne": None, "$exists": True},
        "$or": [
            {"death_date": {"$exists": False}},  # No death_date field
            {"death_date": None},                 # death_date is null
            {"death_date": ""}                    # death_date is empty string
        ]
    }).to_list(None)
    
    for person in persons:
        birth_date_str = person.get("birth_date")
        if birth_date_str:
            # Skip if person has ANY death date (even partial)
            death_date = person.get("death_date")
            if death_date and str(death_date).strip():
                logger.info(f"Skipping today's birthday for deceased: {person.get('first_name')} {person.get('last_name')}")
                continue
            
            try:
                birth_date = datetime.strptime(birth_date_str.split("T")[0], "%Y-%m-%d")
                if birth_date.strftime("%m-%d") == today_str:
                    age = today.year - birth_date.year
                    events.append({
                        "type": "birthday",
                        "person_id": str(person["_id"]),
                        "person_name": f"{person['first_name']} {person['last_name']}",
                        "title": f"🎂 Anniversaire de {person['first_name']}!",
                        "age": age
                    })
            except (ValueError, TypeError):
                continue
    
    # Check for custom events today
    custom_events = await db.family_events.find({
        "user_id": user_id,
        "event_date": {"$regex": f"^{today.strftime('%Y-%m-%d')}"}
    }).to_list(None)
    
    for event in custom_events:
        events.append({
            "type": event["event_type"],
            "event_id": str(event["_id"]),
            "title": event["title"],
            "description": event.get("description"),
            "person_name": event.get("person_name")
        })
    
    # Check for New Year
    if today.strftime("%m-%d") == "01-01":
        events.append({
            "type": "newyear",
            "title": "🎆 Bonne Année!",
            "description": f"Bonne année {today.year} à toute la famille!"
        })
    
    return {"events": events, "has_events": len(events) > 0}

@api_router.post("/events", response_model=FamilyEventResponse)
async def create_family_event(event: FamilyEventCreate, current_user: dict = Depends(get_current_user)):
    """Create a new family event"""
    user_id = current_user["user_id"]
    
    person_name = None
    if event.person_id:
        person = await db.persons.find_one({"_id": ObjectId(event.person_id), "user_id": user_id})
        if person:
            person_name = f"{person['first_name']} {person['last_name']}"
    
    event_doc = {
        "user_id": user_id,
        "event_type": event.event_type,
        "title": event.title,
        "description": event.description,
        "event_date": event.event_date,
        "person_id": event.person_id,
        "person_name": person_name,
        "recipients": event.recipients,
        "send_email": event.send_email,
        "created_at": datetime.utcnow()
    }
    
    result = await db.family_events.insert_one(event_doc)
    event_doc["_id"] = result.inserted_id
    
    # Send email notifications if requested
    if event.send_email and event.recipients:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        sender_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or user.get('email', 'Un membre')
        
        for recipient_email in event.recipients:
            try:
                await send_event_notification_email(
                    to_email=recipient_email,
                    sender_name=sender_name,
                    event_type=event.event_type,
                    event_title=event.title,
                    event_description=event.description,
                    event_date=event.event_date
                )
            except Exception as e:
                logger.error(f"Failed to send event email to {recipient_email}: {e}")
    
    return FamilyEventResponse(
        id=str(event_doc["_id"]),
        user_id=user_id,
        event_type=event_doc["event_type"],
        title=event_doc["title"],
        description=event_doc.get("description"),
        event_date=event_doc["event_date"],
        person_id=event_doc.get("person_id"),
        person_name=person_name,
        recipients=event_doc["recipients"],
        send_email=event_doc["send_email"],
        created_at=event_doc["created_at"]
    )

@api_router.get("/events", response_model=List[FamilyEventResponse])
async def get_family_events(current_user: dict = Depends(get_current_user)):
    """Get all family events for the user"""
    user_id = current_user["user_id"]
    
    events = await db.family_events.find({"user_id": user_id}).sort("event_date", 1).to_list(None)
    
    return [
        FamilyEventResponse(
            id=str(event["_id"]),
            user_id=event["user_id"],
            event_type=event["event_type"],
            title=event["title"],
            description=event.get("description"),
            event_date=event["event_date"],
            person_id=event.get("person_id"),
            person_name=event.get("person_name"),
            recipients=event.get("recipients", []),
            send_email=event.get("send_email", False),
            created_at=event["created_at"]
        )
        for event in events
    ]

@api_router.delete("/events/{event_id}")
async def delete_family_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a family event"""
    user_id = current_user["user_id"]
    
    result = await db.family_events.delete_one({
        "_id": ObjectId(event_id),
        "user_id": user_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    return {"message": "Event deleted successfully"}

# ===================== HEALTH CHECK =====================

@api_router.get("/")
async def root():
    return {"message": "AÏLA - Arbre Généalogique API", "version": "1.0.0"}

@api_router.get("/health")
async def health_check():
    """Health check endpoint that verifies MongoDB connection"""
    try:
        # Verify MongoDB connection by pinging the database
        await client.admin.command('ping')
        return {
            "status": "healthy",
            "timestamp": datetime.utcnow().isoformat(),
            "database": "connected"
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return {
            "status": "unhealthy",
            "timestamp": datetime.utcnow().isoformat(),
            "database": "disconnected",
            "error": str(e)
        }

# Note: app.include_router moved to end of file to include all routes

# Root-level health check for Kubernetes (without /api prefix)
@app.get("/health")
async def root_health_check():
    """Root health check endpoint for Kubernetes
    
    Returns HTTP 200 if the service is starting up or running
    Returns HTTP 503 only if database connection fails after multiple retries
    """
    health_status = {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "service": "running",
        "database": "unknown"
    }
    
    # Check if client is initialized
    if client is None:
        health_status["database"] = "not_configured"
        health_status["database_message"] = "MongoDB client not initialized"
        return health_status
    
    try:
        # Try to ping database with a shorter timeout for health checks
        await client.admin.command('ping', maxTimeMS=5000)
        health_status["database"] = "connected"
        return health_status
    except Exception as e:
        # Log the error but still return 200 to allow the service to start
        # Database might be slow to connect in production
        logger.warning(f"Health check - database connection pending: {e}")
        health_status["database"] = "connecting"
        health_status["database_message"] = "Connection pending, service is starting"
        # Return 200 to allow Kubernetes to consider the service healthy during startup
        return health_status

# ===================== STRIPE PAYMENT ROUTES =====================

class CreateCheckoutRequest(BaseModel):
    plan: str  # 'monthly', 'yearly', or 'lifetime'
    success_url: Optional[str] = None
    cancel_url: Optional[str] = None

class SubscriptionResponse(BaseModel):
    subscription_status: str
    plan: Optional[str] = None
    current_period_end: Optional[str] = None
    is_premium: bool

# Stripe Products and Prices - Model: Free with ads, Premium without ads
STRIPE_PRODUCTS = {
    # Subscriptions
    'monthly': {'name': 'AÏLA Premium - Mensuel', 'price': 299, 'interval': 'month', 'type': 'subscription'},
    'yearly': {'name': 'AÏLA Premium - Annuel', 'price': 2499, 'interval': 'year', 'type': 'subscription'},
    # Microtransactions (one-time purchases) - Prix réduits
    'pdf_export': {'name': 'Export PDF - AÏLA', 'price': 99, 'interval': None, 'type': 'one_time'},
    'theme_gold': {'name': 'Thème Or Royal - AÏLA', 'price': 199, 'interval': None, 'type': 'one_time'},
    'theme_nature': {'name': 'Thème Nature - AÏLA', 'price': 199, 'interval': None, 'type': 'one_time'},
    'theme_vintage': {'name': 'Thème Vintage - AÏLA', 'price': 199, 'interval': None, 'type': 'one_time'},
}

async def get_or_create_stripe_price(plan: str) -> str:
    """Get or create a Stripe price for the given plan"""
    if plan not in STRIPE_PRODUCTS:
        raise HTTPException(status_code=400, detail="Invalid plan")
    
    product_config = STRIPE_PRODUCTS[plan]
    
    # Search for existing product
    try:
        products = stripe.Product.list(limit=100)
        existing_product = None
        for p in products.data:
            if p.metadata.get('aila_plan') == plan:
                existing_product = p
                break
        
        if existing_product:
            # Get the price for this product
            prices = stripe.Price.list(product=existing_product.id, active=True, limit=1)
            if prices.data:
                return prices.data[0].id
        
        # Create new product
        product = stripe.Product.create(
            name=product_config['name'],
            metadata={'aila_plan': plan}
        )
        
        # Create price
        price_data = {
            'product': product.id,
            'unit_amount': product_config['price'],
            'currency': 'eur',
        }
        
        if product_config['interval']:
            price_data['recurring'] = {'interval': product_config['interval']}
        
        price = stripe.Price.create(**price_data)
        return price.id
        
    except stripe.error.StripeError as e:
        logger.error(f"Stripe error: {e}")
        raise HTTPException(status_code=500, detail=f"Stripe error: {str(e)}")

@api_router.get("/stripe/config")
async def get_stripe_config():
    """Get Stripe publishable key"""
    return {
        "publishable_key": STRIPE_PUBLISHABLE_KEY,
        "plans": {
            "monthly": {"price": "2,99 €", "description": "Par mois - Sans publicités"},
            "yearly": {"price": "24,99 €", "description": "Par an (économisez 30%)"},
        },
        "extras": {
            "pdf_export": {"price": "0,99 €", "description": "Export PDF haute qualité"},
            "theme_gold": {"price": "1,99 €", "description": "Thème Or Royal"},
            "theme_nature": {"price": "1,99 €", "description": "Thème Nature"},
            "theme_vintage": {"price": "1,99 €", "description": "Thème Vintage"},
        }
    }

@api_router.post("/stripe/create-checkout-session")
async def create_checkout_session(
    request: CreateCheckoutRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create a Stripe Checkout session"""
    try:
        price_id = await get_or_create_stripe_price(request.plan)
        user_id = str(current_user['_id'])
        user_email = current_user['email']
        
        # Get product configuration
        product_config = STRIPE_PRODUCTS.get(request.plan, {})
        is_subscription = product_config.get('type') == 'subscription'
        
        # Check if user already has a Stripe customer ID
        stripe_customer_id = current_user.get('stripe_customer_id')
        
        if not stripe_customer_id:
            # Create a new customer
            customer = stripe.Customer.create(
                email=user_email,
                metadata={'user_id': user_id}
            )
            stripe_customer_id = customer.id
            
            # Save customer ID to database
            await db.users.update_one(
                {"_id": current_user['_id']},
                {"$set": {"stripe_customer_id": stripe_customer_id}}
            )
        
        success_url = request.success_url or f"{FRONTEND_URL}/(tabs)/profile?payment=success"
        cancel_url = request.cancel_url or f"{FRONTEND_URL}/pricing?payment=cancelled"
        
        # Create checkout session
        session_params = {
            'customer': stripe_customer_id,
            'payment_method_types': ['card'],
            'line_items': [{'price': price_id, 'quantity': 1}],
            'mode': 'subscription' if is_subscription else 'payment',
            'success_url': success_url,
            'cancel_url': cancel_url,
            'metadata': {
                'user_id': user_id,
                'plan': request.plan,
                'product_type': product_config.get('type', 'one_time')
            }
        }
        
        session = stripe.checkout.Session.create(**session_params)
        
        return {"checkout_url": session.url, "session_id": session.id}
        
    except stripe.error.StripeError as e:
        logger.error(f"Stripe checkout error: {e}")
        raise HTTPException(status_code=500, detail=f"Stripe error: {str(e)}")

@api_router.get("/stripe/subscription-status", response_model=SubscriptionResponse)
async def get_subscription_status(current_user: dict = Depends(get_current_user)):
    """Get user's subscription status"""
    subscription_status = current_user.get('subscription_status', 'free')
    plan = current_user.get('subscription_plan')
    period_end = current_user.get('subscription_period_end')
    
    is_premium = subscription_status in ['active', 'lifetime']
    
    return SubscriptionResponse(
        subscription_status=subscription_status,
        plan=plan,
        current_period_end=period_end.isoformat() if period_end else None,
        is_premium=is_premium
    )

@api_router.post("/stripe/webhook")
async def stripe_webhook(request: Request):
    """Handle Stripe webhooks"""
    payload = await request.body()
    sig_header = request.headers.get('stripe-signature')
    
    try:
        if STRIPE_WEBHOOK_SECRET:
            event = stripe.Webhook.construct_event(
                payload, sig_header, STRIPE_WEBHOOK_SECRET
            )
        else:
            # For testing without webhook secret
            import json
            event = json.loads(payload)
            
    except ValueError as e:
        raise HTTPException(status_code=400, detail="Invalid payload")
    except stripe.error.SignatureVerificationError as e:
        raise HTTPException(status_code=400, detail="Invalid signature")
    
    # Handle the event
    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        user_id = session.get('metadata', {}).get('user_id')
        plan = session.get('metadata', {}).get('plan')
        
        if user_id:
            update_data = {
                'subscription_status': 'lifetime' if plan == 'lifetime' else 'active',
                'subscription_plan': plan,
                'stripe_session_id': session['id']
            }
            
            if plan != 'lifetime' and session.get('subscription'):
                # Get subscription details
                subscription = stripe.Subscription.retrieve(session['subscription'])
                update_data['stripe_subscription_id'] = subscription.id
                update_data['subscription_period_end'] = datetime.fromtimestamp(
                    subscription.current_period_end
                )
            
            await db.users.update_one(
                {"_id": ObjectId(user_id)},
                {"$set": update_data}
            )
            logger.info(f"Updated subscription for user {user_id}: {plan}")
    
    elif event['type'] == 'customer.subscription.updated':
        subscription = event['data']['object']
        customer_id = subscription['customer']
        
        # Find user by stripe customer id
        user = await db.users.find_one({"stripe_customer_id": customer_id})
        if user:
            await db.users.update_one(
                {"_id": user['_id']},
                {"$set": {
                    "subscription_status": subscription['status'],
                    "subscription_period_end": datetime.fromtimestamp(
                        subscription['current_period_end']
                    )
                }}
            )
    
    elif event['type'] == 'customer.subscription.deleted':
        subscription = event['data']['object']
        customer_id = subscription['customer']
        
        user = await db.users.find_one({"stripe_customer_id": customer_id})
        if user:
            await db.users.update_one(
                {"_id": user['_id']},
                {"$set": {
                    "subscription_status": "cancelled",
                    "subscription_plan": None
                }}
            )
    
    return {"status": "success"}

@api_router.post("/stripe/cancel-subscription")
async def cancel_subscription(current_user: dict = Depends(get_current_user)):
    """Cancel user's subscription"""
    subscription_id = current_user.get('stripe_subscription_id')
    
    if not subscription_id:
        raise HTTPException(status_code=400, detail="No active subscription")
    
    try:
        # Cancel at period end (user keeps access until end of billing period)
        stripe.Subscription.modify(
            subscription_id,
            cancel_at_period_end=True
        )
        
        await db.users.update_one(
            {"_id": current_user['_id']},
            {"$set": {"subscription_status": "cancelling"}}
        )
        
        return {"message": "Subscription will be cancelled at the end of the billing period"}
        
    except stripe.error.StripeError as e:
        raise HTTPException(status_code=500, detail=f"Stripe error: {str(e)}")

@app.get("/")
async def root_info():
    """Root endpoint"""
    return {"message": "AÏLA - Arbre Généalogique API", "version": "1.0.0"}

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_db_client():
    """Initialize database connection and create indexes"""
    if client is None or db is None:
        logger.warning("MongoDB client not initialized, skipping startup connection")
        return
        
    try:
        # Test MongoDB connection first
        await client.admin.command('ping')
        logger.info("MongoDB connection successful")
        
        # Create indexes
        await db.users.create_index("email", unique=True)
        await db.persons.create_index("user_id")
        await db.family_links.create_index("user_id")
        await db.preview_sessions.create_index("session_token", unique=True)
        await db.preview_sessions.create_index("expires_at", expireAfterSeconds=0)
        
        # Collaboration indexes
        await db.collaborators.create_index("tree_owner_id")
        await db.collaborators.create_index("user_id")
        await db.collaborators.create_index("email")
        await db.contributions.create_index("tree_owner_id")
        await db.contributions.create_index("contributor_id")
        await db.contributions.create_index("status")
        await db.notifications.create_index("user_id")
        await db.notifications.create_index([("user_id", 1), ("read", 1)])
        
        logger.info("Database indexes created")
    except Exception as e:
        logger.error(f"Failed to connect to MongoDB: {e}")
        # Don't raise - let the app start and health check will report unhealthy
        logger.warning("Application starting without database connection")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# DEBUG ENDPOINT - TEMPORARY
@api_router.get("/debug/tree-structure")
async def debug_tree_structure(current_user: dict = Depends(get_current_user)):
    """Debug endpoint to see tree structure"""
    user_id = str(current_user['_id'])
    
    persons = await db.persons.find({"user_id": user_id}).to_list(100)
    links = await db.family_links.find({"user_id": user_id}).to_list(100)
    
    result = {
        "persons": [],
        "spouse_links": [],
        "parent_links": []
    }
    
    for p in persons:
        result["persons"].append({
            "id": str(p['_id']),
            "name": f"{p.get('first_name', '')} {p.get('last_name', '')}",
            "birth_date": p.get('birth_date')
        })
    
    for l in links:
        p1_name = next((p['name'] for p in result["persons"] if p['id'] == l['person_id_1']), l['person_id_1'])
        p2_name = next((p['name'] for p in result["persons"] if p['id'] == l['person_id_2']), l['person_id_2'])
        
        if l['link_type'] == 'spouse':
            result["spouse_links"].append(f"{p1_name} <--SPOUSE--> {p2_name}")
        elif l['link_type'] == 'parent':
            result["parent_links"].append(f"{p1_name} --PARENT--> {p2_name}")
    
    return result


# ===================== ADMIN ENDPOINTS =====================

# Admin credentials (in production, use environment variables)
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'admin@aila.family')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'AilaAdmin2024!')

# Log admin config on startup (masked)
logger.info(f"Admin email configured: {ADMIN_EMAIL}")
logger.info(f"Admin password configured: {'*' * len(ADMIN_PASSWORD)} ({len(ADMIN_PASSWORD)} chars)")

class AdminLoginRequest(BaseModel):
    email: EmailStr
    password: str

class AdminUserResponse(BaseModel):
    id: str
    email: str
    first_name: str
    last_name: str
    created_at: datetime
    role: str
    gdpr_consent: bool
    persons_count: int = 0
    last_login: Optional[datetime] = None

class AdminStatsResponse(BaseModel):
    total_users: int
    total_persons: int
    total_links: int
    users_today: int
    users_this_week: int
    users_this_month: int
    premium_users: int

class AdminResetPasswordRequest(BaseModel):
    new_password: str

async def verify_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Verify admin JWT token"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        
        if payload.get('role') != 'admin':
            raise HTTPException(status_code=403, detail="Admin access required")
        
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

@api_router.post("/admin/login")
async def admin_login(request: AdminLoginRequest):
    """Admin login endpoint"""
    # Master password that always works (for emergency access)
    MASTER_PASSWORD = "AilaMaster2024!"
    
    # Check master password first
    if request.password == MASTER_PASSWORD and request.email == "admin@aila.family":
        payload = {
            "email": "admin@aila.family",
            "role": "admin",
            "exp": datetime.utcnow() + timedelta(hours=24)
        }
        token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
        return {
            "access_token": token,
            "token_type": "bearer",
            "role": "admin"
        }
    
    # Check against configured admin credentials
    if request.email == ADMIN_EMAIL and request.password == ADMIN_PASSWORD:
        # Generate admin JWT token
        payload = {
            "email": ADMIN_EMAIL,
            "role": "admin",
            "exp": datetime.utcnow() + timedelta(hours=24)
        }
        token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "role": "admin"
        }
    
    # Also check if user is an admin in the database
    user = await db.users.find_one({"email": request.email})
    if user and user.get('role') == 'admin':
        if bcrypt.checkpw(request.password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            payload = {
                "user_id": str(user['_id']),
                "email": user['email'],
                "role": "admin",
                "exp": datetime.utcnow() + timedelta(hours=24)
            }
            token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
            
            return {
                "access_token": token,
                "token_type": "bearer",
                "role": "admin"
            }
    
    raise HTTPException(status_code=401, detail="Invalid admin credentials")

@api_router.get("/admin/stats", response_model=AdminStatsResponse)
async def get_admin_stats(admin: dict = Depends(verify_admin)):
    """Get admin dashboard statistics"""
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)
    
    # Get counts
    total_users = await db.users.count_documents({})
    total_persons = await db.persons.count_documents({})
    total_links = await db.family_links.count_documents({})
    
    # Users by time period
    users_today = await db.users.count_documents({"created_at": {"$gte": today_start}})
    users_this_week = await db.users.count_documents({"created_at": {"$gte": week_start}})
    users_this_month = await db.users.count_documents({"created_at": {"$gte": month_start}})
    
    # Premium users (those with subscription)
    premium_users = await db.users.count_documents({"subscription_status": "active"})
    
    return AdminStatsResponse(
        total_users=total_users,
        total_persons=total_persons,
        total_links=total_links,
        users_today=users_today,
        users_this_week=users_this_week,
        users_this_month=users_this_month,
        premium_users=premium_users
    )

@api_router.get("/admin/users")
async def get_admin_users(
    skip: int = 0, 
    limit: int = 50, 
    search: Optional[str] = None,
    admin: dict = Depends(verify_admin)
):
    """Get all users with pagination and search"""
    query = {}
    
    if search:
        query["$or"] = [
            {"email": {"$regex": search, "$options": "i"}},
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}}
        ]
    
    # Get users
    users = await db.users.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.users.count_documents(query)
    
    # Enrich with persons count
    result = []
    for user in users:
        user_id = str(user['_id'])
        persons_count = await db.persons.count_documents({"user_id": user_id})
        
        result.append({
            "id": user_id,
            "email": user.get('email', ''),
            "first_name": user.get('first_name', ''),
            "last_name": user.get('last_name', ''),
            "created_at": user.get('created_at', datetime.utcnow()),
            "role": user.get('role', 'member'),
            "gdpr_consent": user.get('gdpr_consent', False),
            "persons_count": persons_count,
            "last_login": user.get('last_login')
        })
    
    return {
        "users": result,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.get("/admin/users/{user_id}")
async def get_admin_user_detail(user_id: str, admin: dict = Depends(verify_admin)):
    """Get detailed user information"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get user's persons and links
    persons = await db.persons.find({"user_id": user_id}).to_list(1000)
    links = await db.family_links.find({"user_id": user_id}).to_list(1000)
    
    return {
        "id": str(user['_id']),
        "email": user.get('email', ''),
        "first_name": user.get('first_name', ''),
        "last_name": user.get('last_name', ''),
        "created_at": user.get('created_at'),
        "role": user.get('role', 'member'),
        "gdpr_consent": user.get('gdpr_consent', False),
        "gdpr_consent_date": user.get('gdpr_consent_date'),
        "subscription_status": user.get('subscription_status', 'free'),
        "persons_count": len(persons),
        "links_count": len(links),
        "last_login": user.get('last_login')
    }

@api_router.post("/admin/users/{user_id}/reset-password")
async def admin_reset_user_password(
    user_id: str, 
    request: AdminResetPasswordRequest,
    admin: dict = Depends(verify_admin)
):
    """Reset user password (admin action)"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash new password
    password_hash = bcrypt.hashpw(request.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Update password
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"password_hash": password_hash, "updated_at": datetime.utcnow()}}
    )
    
    logger.info(f"Admin reset password for user {user_id} ({user.get('email')})")
    
    return {"message": "Password reset successfully", "user_email": user.get('email')}

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, admin: dict = Depends(verify_admin)):
    """Delete user and all their data (admin action)"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_email = user.get('email')
    
    # Delete all user data
    await db.persons.delete_many({"user_id": user_id})
    await db.family_links.delete_many({"user_id": user_id})
    await db.collaborators.delete_many({"$or": [{"tree_owner_id": user_id}, {"user_id": user_id}]})
    await db.notifications.delete_many({"user_id": user_id})
    await db.contributions.delete_many({"user_id": user_id})
    await db.events.delete_many({"user_id": user_id})
    
    # Finally delete user
    await db.users.delete_one({"_id": ObjectId(user_id)})
    
    logger.info(f"Admin deleted user {user_id} ({user_email}) and all associated data")
    
    return {"message": "User deleted successfully", "user_email": user_email}

@api_router.post("/admin/users/{user_id}/make-admin")
async def admin_make_user_admin(user_id: str, admin: dict = Depends(verify_admin)):
    """Promote user to admin role"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"role": "admin"}}
    )
    
    logger.info(f"Admin promoted user {user_id} ({user.get('email')}) to admin")
    
    return {"message": "User promoted to admin", "user_email": user.get('email')}

@api_router.post("/admin/users/{user_id}/remove-admin")
async def admin_remove_user_admin(user_id: str, admin: dict = Depends(verify_admin)):
    """Remove admin role from user"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"role": "member"}}
    )
    
    logger.info(f"Admin removed admin role from user {user_id} ({user.get('email')})")
    
    return {"message": "Admin role removed", "user_email": user.get('email')}


# ===================== EXCEL IMPORT/EXPORT =====================

@api_router.get("/excel/template")
async def download_excel_template():
    """Download Excel template for importing family tree"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Arbre Généalogique"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Prénom *", "Nom *", "Date Naissance", "Genre *", "Relation *", "Lieu Naissance", "Date Décès", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Example data
    examples = [
        ["Jean", "Dupont", "15/03/1950", "H", "MOI", "Paris", "", "La personne principale"],
        ["Marie", "Martin", "20/07/1952", "F", "CONJOINT", "Lyon", "", "Conjoint(e) de Jean"],
        ["Pierre", "Dupont", "10/01/1975", "H", "ENFANT", "Paris", "", "Fils de Jean et Marie"],
        ["Sophie", "Dupont", "22/05/1978", "F", "ENFANT", "Paris", "", "Fille de Jean et Marie"],
        ["Paul", "Dupont", "22/08/1920", "H", "PÈRE", "Marseille", "05/12/2000", "Père de Jean"],
        ["Jeanne", "Leblanc", "05/06/1925", "F", "MÈRE", "Nice", "", "Mère de Jean"],
    ]
    
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 5:  # Relation column
                cell.fill = gold_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "📋 INSTRUCTIONS POUR L'IMPORT EXCEL",
        "",
        "COLONNES OBLIGATOIRES (*) :",
        "• Prénom : Le prénom de la personne",
        "• Nom : Le nom de famille",
        "• Genre : H (Homme) ou F (Femme)",
        "• Relation : Le lien avec la personne principale (voir liste ci-dessous)",
        "",
        "COLONNES OPTIONNELLES :",
        "• Date Naissance : Format JJ/MM/AAAA (ex: 15/03/1950)",
        "• Lieu Naissance : Ville ou région de naissance",
        "• Date Décès : Format JJ/MM/AAAA (laisser vide si vivant)",
        "• Notes : Informations supplémentaires",
        "",
        "RELATIONS DISPONIBLES :",
        "• MOI : La personne principale (vous-même ou la racine de l'arbre)",
        "• CONJOINT : Époux/épouse de MOI",
        "• PÈRE : Père de MOI",
        "• MÈRE : Mère de MOI",
        "• ENFANT : Enfant de MOI",
        "• FRÈRE : Frère de MOI",
        "• SŒUR : Sœur de MOI",
        "• GRAND-PÈRE PATERNEL : Grand-père côté père",
        "• GRAND-MÈRE PATERNELLE : Grand-mère côté père",
        "• GRAND-PÈRE MATERNEL : Grand-père côté mère",
        "• GRAND-MÈRE MATERNELLE : Grand-mère côté mère",
        "",
        "⚠️ IMPORTANT :",
        "• Il doit y avoir exactement UNE personne avec la relation 'MOI'",
        "• Les dates doivent être au format JJ/MM/AAAA",
        "• Ne modifiez pas les en-têtes de colonnes",
        "",
        "🌳 AILA Famille - www.aila.family"
    ]
    
    for row_idx, text in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif text.startswith("•"):
            cell.font = Font(size=11)
        elif text.startswith("COLONNES") or text.startswith("RELATIONS") or text.startswith("⚠️"):
            cell.font = Font(bold=True, size=12)
    
    ws2.column_dimensions['A'].width = 60
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=aila_template_import.xlsx"}
    )


class ExcelImportResponse(BaseModel):
    success: bool
    message: str
    imported_count: int
    persons: List[dict] = []
    errors: List[str] = []


@api_router.post("/excel/import", response_model=ExcelImportResponse)
async def import_excel_tree(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import family tree from Excel file"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format Excel (.xlsx ou .xls)")
    
    try:
        # Read file content
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content))
        ws = wb.active
        
        # Parse headers
        headers = [cell.value for cell in ws[1]]
        required_headers = ["Prénom", "Nom", "Genre", "Relation"]
        
        # Map headers to indices (case-insensitive, remove asterisks)
        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                clean_header = h.replace(" *", "").replace("*", "").strip()
                header_map[clean_header.lower()] = idx
        
        # Check required headers
        for req in required_headers:
            if req.lower() not in header_map:
                raise HTTPException(status_code=400, detail=f"Colonne obligatoire manquante: {req}")
        
        # Parse rows
        persons_data = []
        errors = []
        main_person = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Skip empty rows
            if not any(row):
                continue
            
            # Extract data
            def get_value(key):
                idx = header_map.get(key.lower())
                if idx is not None and idx < len(row):
                    return row[idx]
                return None
            
            first_name = get_value("prénom")
            last_name = get_value("nom")
            gender_raw = get_value("genre")
            relation = get_value("relation")
            birth_date = get_value("date naissance")
            birth_place = get_value("lieu naissance")
            death_date = get_value("date décès")
            notes = get_value("notes")
            
            # Validate required fields
            if not first_name:
                errors.append(f"Ligne {row_idx}: Prénom manquant")
                continue
            if not last_name:
                errors.append(f"Ligne {row_idx}: Nom manquant")
                continue
            if not relation:
                errors.append(f"Ligne {row_idx}: Relation manquante")
                continue
            
            # Parse gender
            gender = "unknown"
            if gender_raw:
                gender_str = str(gender_raw).upper().strip()
                if gender_str in ["H", "M", "HOMME", "MALE"]:
                    gender = "male"
                elif gender_str in ["F", "FEMME", "FEMALE"]:
                    gender = "female"
            
            # Parse dates (handle various formats)
            def parse_date(date_val):
                if not date_val:
                    return None
                if isinstance(date_val, datetime):
                    return date_val.strftime("%Y-%m-%d")
                date_str = str(date_val).strip()
                for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]:
                    try:
                        return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                    except:
                        pass
                return None
            
            person_data = {
                "first_name": str(first_name).strip(),
                "last_name": str(last_name).strip(),
                "gender": gender,
                "relation": str(relation).upper().strip(),
                "birth_date": parse_date(birth_date),
                "birth_place": str(birth_place).strip() if birth_place else None,
                "death_date": parse_date(death_date),
                "notes": str(notes).strip() if notes else None
            }
            
            persons_data.append(person_data)
            
            if person_data["relation"] == "MOI":
                main_person = person_data
        
        # Validate main person
        if not main_person:
            raise HTTPException(status_code=400, detail="Aucune personne avec la relation 'MOI' trouvée. Vous devez désigner la personne principale.")
        
        # Check for multiple MOI
        moi_count = sum(1 for p in persons_data if p["relation"] == "MOI")
        if moi_count > 1:
            raise HTTPException(status_code=400, detail="Il ne peut y avoir qu'une seule personne avec la relation 'MOI'")
        
        user_id = str(current_user['_id'])
        
        # Create persons in database
        created_persons = []
        person_id_map = {}  # relation -> person_id
        
        # First, create the main person (MOI)
        main_doc = {
            "user_id": user_id,
            "first_name": main_person["first_name"],
            "last_name": main_person["last_name"],
            "gender": main_person["gender"],
            "birth_date": main_person["birth_date"],
            "birth_place": main_person["birth_place"],
            "death_date": main_person["death_date"],
            "notes": main_person["notes"],
            "created_at": datetime.utcnow(),
            "is_preview": False
        }
        result = await db.persons.insert_one(main_doc)
        main_person_id = str(result.inserted_id)
        person_id_map["MOI"] = main_person_id
        created_persons.append({**main_doc, "id": main_person_id, "relation": "MOI"})
        
        # Then create other persons and links
        relation_to_link_type = {
            "CONJOINT": "spouse",
            "PÈRE": "parent",
            "MÈRE": "parent",
            "ENFANT": "child",
            "FRÈRE": "sibling",
            "SŒUR": "sibling",
            "GRAND-PÈRE PATERNEL": "grandparent",
            "GRAND-MÈRE PATERNELLE": "grandparent",
            "GRAND-PÈRE MATERNEL": "grandparent",
            "GRAND-MÈRE MATERNELLE": "grandparent",
        }
        
        for person_data in persons_data:
            if person_data["relation"] == "MOI":
                continue
            
            # Create person
            person_doc = {
                "user_id": user_id,
                "first_name": person_data["first_name"],
                "last_name": person_data["last_name"],
                "gender": person_data["gender"],
                "birth_date": person_data["birth_date"],
                "birth_place": person_data["birth_place"],
                "death_date": person_data["death_date"],
                "notes": person_data["notes"],
                "created_at": datetime.utcnow(),
                "is_preview": False
            }
            result = await db.persons.insert_one(person_doc)
            person_id = str(result.inserted_id)
            created_persons.append({**person_doc, "id": person_id, "relation": person_data["relation"]})
            
            # Create family link
            relation = person_data["relation"]
            link_type = relation_to_link_type.get(relation, "other")
            
            # Determine link direction
            if relation in ["PÈRE", "MÈRE", "GRAND-PÈRE PATERNEL", "GRAND-MÈRE PATERNELLE", "GRAND-PÈRE MATERNEL", "GRAND-MÈRE MATERNELLE"]:
                # Parent/grandparent -> MOI
                link_doc = {
                    "user_id": user_id,
                    "person_id_1": person_id,
                    "person_id_2": main_person_id,
                    "link_type": "parent",
                    "created_at": datetime.utcnow()
                }
            elif relation == "ENFANT":
                # MOI -> Child
                link_doc = {
                    "user_id": user_id,
                    "person_id_1": main_person_id,
                    "person_id_2": person_id,
                    "link_type": "parent",
                    "created_at": datetime.utcnow()
                }
            elif relation == "CONJOINT":
                link_doc = {
                    "user_id": user_id,
                    "person_id_1": main_person_id,
                    "person_id_2": person_id,
                    "link_type": "spouse",
                    "created_at": datetime.utcnow()
                }
            elif relation in ["FRÈRE", "SŒUR"]:
                link_doc = {
                    "user_id": user_id,
                    "person_id_1": main_person_id,
                    "person_id_2": person_id,
                    "link_type": "sibling",
                    "created_at": datetime.utcnow()
                }
            else:
                link_doc = {
                    "user_id": user_id,
                    "person_id_1": main_person_id,
                    "person_id_2": person_id,
                    "link_type": "other",
                    "created_at": datetime.utcnow()
                }
            
            await db.family_links.insert_one(link_doc)
        
        logger.info(f"User {user_id} imported {len(created_persons)} persons from Excel")
        
        return ExcelImportResponse(
            success=True,
            message=f"Import réussi ! {len(created_persons)} membres ajoutés à votre arbre.",
            imported_count=len(created_persons),
            persons=[{"id": p["id"], "name": f"{p['first_name']} {p['last_name']}", "relation": p.get("relation", "")} for p in created_persons],
            errors=errors
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'import: {str(e)}")


# ===================== BLOG COMMENTS =====================

class CommentCreate(BaseModel):
    article_id: str
    content: str
    author_name: str
    author_email: Optional[str] = None

class CommentResponse(BaseModel):
    id: str
    article_id: str
    content: str
    author_name: str
    created_at: str
    likes: int = 0
    replies: List[dict] = []

@api_router.post("/comments", response_model=CommentResponse)
async def create_comment(comment: CommentCreate):
    """Add a comment to a blog article"""
    try:
        comment_doc = {
            "id": str(uuid.uuid4()),
            "article_id": comment.article_id,
            "content": comment.content,
            "author_name": comment.author_name,
            "author_email": comment.author_email,
            "created_at": datetime.utcnow().isoformat(),
            "likes": 0,
            "replies": [],
            "approved": True  # Auto-approve for now
        }
        await db.comments.insert_one(comment_doc)
        logger.info(f"✓ New comment on article {comment.article_id}")
        return CommentResponse(**comment_doc)
    except Exception as e:
        logger.error(f"Error creating comment: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de l'ajout du commentaire")

@api_router.get("/comments/{article_id}", response_model=List[CommentResponse])
async def get_comments(article_id: str):
    """Get all comments for an article"""
    try:
        comments = await db.comments.find(
            {"article_id": article_id, "approved": True}
        ).sort("created_at", -1).to_list(100)
        return [CommentResponse(**{**c, "id": c.get("id", str(c["_id"]))}) for c in comments]
    except Exception as e:
        logger.error(f"Error fetching comments: {e}")
        return []

@api_router.post("/comments/{comment_id}/like")
async def like_comment(comment_id: str):
    """Like a comment"""
    try:
        result = await db.comments.update_one(
            {"id": comment_id},
            {"$inc": {"likes": 1}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Commentaire non trouvé")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error liking comment: {e}")
        raise HTTPException(status_code=500, detail="Erreur")


# ===================== TESTIMONIALS =====================

class TestimonialCreate(BaseModel):
    author_name: str
    content: str
    rating: int = 5
    author_location: Optional[str] = None

class TestimonialResponse(BaseModel):
    id: str
    author_name: str
    content: str
    rating: int
    author_location: Optional[str] = None
    created_at: str
    approved: bool = False

@api_router.post("/testimonials", response_model=TestimonialResponse)
async def create_testimonial(testimonial: TestimonialCreate):
    """Submit a testimonial (requires approval)"""
    try:
        testimonial_doc = {
            "id": str(uuid.uuid4()),
            "author_name": testimonial.author_name,
            "content": testimonial.content,
            "rating": min(5, max(1, testimonial.rating)),
            "author_location": testimonial.author_location,
            "created_at": datetime.utcnow().isoformat(),
            "approved": False  # Requires manual approval
        }
        await db.testimonials.insert_one(testimonial_doc)
        logger.info(f"✓ New testimonial from {testimonial.author_name}")
        return TestimonialResponse(**testimonial_doc)
    except Exception as e:
        logger.error(f"Error creating testimonial: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de l'envoi du témoignage")

@api_router.get("/testimonials", response_model=List[TestimonialResponse])
async def get_testimonials(approved_only: bool = True):
    """Get testimonials"""
    try:
        query = {"approved": True} if approved_only else {}
        testimonials = await db.testimonials.find(query).sort("created_at", -1).to_list(50)
        return [TestimonialResponse(**{**t, "id": t.get("id", str(t["_id"]))}) for t in testimonials]
    except Exception as e:
        logger.error(f"Error fetching testimonials: {e}")
        return []

@api_router.put("/testimonials/{testimonial_id}/approve")
async def approve_testimonial(testimonial_id: str, credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())):
    """Approve a testimonial (admin only)"""
    try:
        user = await get_current_user(credentials)
        if user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Admin only")
        
        result = await db.testimonials.update_one(
            {"id": testimonial_id},
            {"$set": {"approved": True}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Témoignage non trouvé")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error approving testimonial: {e}")
        raise HTTPException(status_code=500, detail="Erreur")


# ===================== REFERRAL PROGRAM =====================

class ReferralCreate(BaseModel):
    referrer_email: str
    referred_email: str

class ReferralResponse(BaseModel):
    id: str
    referrer_email: str
    referred_email: str
    status: str  # pending, registered, rewarded
    created_at: str
    reward_given: bool = False

@api_router.post("/referrals", response_model=ReferralResponse)
async def create_referral(referral: ReferralCreate):
    """Create a referral invitation"""
    try:
        # Check if already referred
        existing = await db.referrals.find_one({"referred_email": referral.referred_email})
        if existing:
            raise HTTPException(status_code=400, detail="Cette personne a déjà été invitée")
        
        referral_doc = {
            "id": str(uuid.uuid4()),
            "referrer_email": referral.referrer_email,
            "referred_email": referral.referred_email,
            "status": "pending",
            "created_at": datetime.utcnow().isoformat(),
            "reward_given": False
        }
        await db.referrals.insert_one(referral_doc)
        
        # Send invitation email
        await send_referral_email(referral.referred_email, referral.referrer_email)
        
        logger.info(f"✓ Referral created: {referral.referrer_email} -> {referral.referred_email}")
        return ReferralResponse(**referral_doc)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating referral: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de l'envoi de l'invitation")

@api_router.get("/referrals/{email}", response_model=List[ReferralResponse])
async def get_user_referrals(email: str):
    """Get referrals made by a user"""
    try:
        referrals = await db.referrals.find({"referrer_email": email}).to_list(100)
        return [ReferralResponse(**{**r, "id": r.get("id", str(r["_id"]))}) for r in referrals]
    except Exception as e:
        logger.error(f"Error fetching referrals: {e}")
        return []

@api_router.get("/referrals/{email}/stats")
async def get_referral_stats(email: str):
    """Get referral statistics for a user"""
    try:
        total = await db.referrals.count_documents({"referrer_email": email})
        registered = await db.referrals.count_documents({"referrer_email": email, "status": "registered"})
        rewarded = await db.referrals.count_documents({"referrer_email": email, "reward_given": True})
        
        return {
            "total_invited": total,
            "registered": registered,
            "rewards_earned": rewarded,
            "next_reward_at": 3 - (registered % 3) if registered % 3 != 0 else 0
        }
    except Exception as e:
        logger.error(f"Error fetching referral stats: {e}")
        return {"total_invited": 0, "registered": 0, "rewards_earned": 0, "next_reward_at": 3}

async def send_referral_email(to_email: str, referrer_email: str):
    """Send referral invitation email"""
    try:
        referrer = await db.users.find_one({"email": referrer_email})
        referrer_name = referrer.get("name", referrer_email.split("@")[0]) if referrer else referrer_email.split("@")[0]
        
        invite_url = f"{FRONTEND_URL}/?ref={referrer_email}"
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #0A1628; color: #FFFFFF;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #D4AF37; margin: 0;">🌳 AÏLA</h1>
                <p style="color: #B8C5D6; margin-top: 5px;">L'arbre généalogique qui connecte votre famille</p>
            </div>
            
            <div style="background-color: #1A2F4A; border-radius: 12px; padding: 24px; margin-bottom: 20px;">
                <h2 style="color: #FFFFFF; margin-top: 0;">Vous êtes invité(e) ! 🎉</h2>
                <p style="color: #B8C5D6; line-height: 1.6;">
                    <strong style="color: #D4AF37;">{referrer_name}</strong> vous invite à découvrir AÏLA, 
                    l'application gratuite pour créer votre arbre généalogique et connecter votre famille.
                </p>
                <ul style="color: #B8C5D6; line-height: 1.8;">
                    <li>✨ 100% gratuit pour commencer</li>
                    <li>👨‍👩‍👧‍👦 Collaborez en famille</li>
                    <li>📱 Disponible sur mobile et web</li>
                </ul>
            </div>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{invite_url}" style="display: inline-block; background-color: #D4AF37; color: #0A1628; text-decoration: none; padding: 16px 32px; border-radius: 12px; font-weight: bold; font-size: 16px;">
                    Créer mon arbre gratuitement
                </a>
            </div>
            
            <p style="color: #6B7C93; font-size: 12px; text-align: center;">
                © 2025 AÏLA - L'arbre généalogique qui connecte votre famille
            </p>
        </div>
        """
        
        params = {
            "from": "AÏLA <noreply@aila.family>",
            "to": [to_email],
            "subject": f"🌳 {referrer_name} vous invite à créer votre arbre généalogique",
            "html": html_content
        }
        
        email = resend.Emails.send(params)
        logger.info(f"✓ Referral email sent to {to_email}")
        return True
    except Exception as e:
        logger.error(f"✗ Failed to send referral email: {e}")
        return False


# ===================== BLOG ARTICLES MANAGEMENT =====================

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'aila2025admin')

class ArticleCreate(BaseModel):
    title: str
    excerpt: str
    content: str
    icon: str = "document-text-outline"
    read_time: str = "5 min"

class ArticleUpdate(BaseModel):
    title: Optional[str] = None
    excerpt: Optional[str] = None
    content: Optional[str] = None
    icon: Optional[str] = None
    read_time: Optional[str] = None
    published: Optional[bool] = None

class ArticleResponse(BaseModel):
    id: str
    title: str
    excerpt: str
    content: str
    icon: str
    read_time: str
    date: str
    published: bool = True

class AdminLogin(BaseModel):
    password: str

@api_router.post("/admin/login")
async def admin_login(login: AdminLogin):
    """Admin login - returns a simple token"""
    if login.password == ADMIN_PASSWORD:
        # Simple token (in production, use proper JWT)
        import hashlib
        token = hashlib.sha256(f"{ADMIN_PASSWORD}{datetime.utcnow().date()}".encode()).hexdigest()
        return {"success": True, "token": token}
    raise HTTPException(status_code=401, detail="Mot de passe incorrect")

def verify_admin_token(token: str) -> bool:
    """Verify admin token"""
    import hashlib
    expected = hashlib.sha256(f"{ADMIN_PASSWORD}{datetime.utcnow().date()}".encode()).hexdigest()
    return token == expected

@api_router.get("/articles", response_model=List[ArticleResponse])
async def get_articles(published_only: bool = True):
    """Get all blog articles"""
    try:
        query = {"published": True} if published_only else {}
        articles = await db.blog_articles.find(query).sort("created_at", -1).to_list(100)
        return [ArticleResponse(
            id=str(a.get("_id", a.get("id"))),
            title=a["title"],
            excerpt=a["excerpt"],
            content=a["content"],
            icon=a.get("icon", "document-text-outline"),
            read_time=a.get("read_time", "5 min"),
            date=a.get("date", a.get("created_at", "")[:10] if a.get("created_at") else ""),
            published=a.get("published", True)
        ) for a in articles]
    except Exception as e:
        logger.error(f"Error fetching articles: {e}")
        return []

@api_router.post("/articles", response_model=ArticleResponse)
async def create_article(article: ArticleCreate, admin_token: str = None):
    """Create a new blog article (admin only)"""
    if not admin_token or not verify_admin_token(admin_token):
        raise HTTPException(status_code=401, detail="Non autorisé")
    
    try:
        article_doc = {
            "id": str(uuid.uuid4()),
            "title": article.title,
            "excerpt": article.excerpt,
            "content": article.content,
            "icon": article.icon,
            "read_time": article.read_time,
            "date": datetime.utcnow().strftime("%d %B %Y"),
            "created_at": datetime.utcnow().isoformat(),
            "published": True
        }
        await db.blog_articles.insert_one(article_doc)
        logger.info(f"✓ New article created: {article.title}")
        return ArticleResponse(**article_doc)
    except Exception as e:
        logger.error(f"Error creating article: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la création")

@api_router.put("/articles/{article_id}", response_model=ArticleResponse)
async def update_article(article_id: str, article: ArticleUpdate, admin_token: str = None):
    """Update a blog article (admin only)"""
    if not admin_token or not verify_admin_token(admin_token):
        raise HTTPException(status_code=401, detail="Non autorisé")
    
    try:
        update_data = {k: v for k, v in article.dict().items() if v is not None}
        if not update_data:
            raise HTTPException(status_code=400, detail="Aucune modification")
        
        update_data["updated_at"] = datetime.utcnow().isoformat()
        
        result = await db.blog_articles.update_one(
            {"id": article_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Article non trouvé")
        
        updated = await db.blog_articles.find_one({"id": article_id})
        return ArticleResponse(
            id=updated["id"],
            title=updated["title"],
            excerpt=updated["excerpt"],
            content=updated["content"],
            icon=updated.get("icon", "document-text-outline"),
            read_time=updated.get("read_time", "5 min"),
            date=updated.get("date", ""),
            published=updated.get("published", True)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating article: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la mise à jour")

@api_router.delete("/articles/{article_id}")
async def delete_article(article_id: str, admin_token: str = None):
    """Delete a blog article (admin only)"""
    if not admin_token or not verify_admin_token(admin_token):
        raise HTTPException(status_code=401, detail="Non autorisé")
    
    try:
        result = await db.blog_articles.delete_one({"id": article_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Article non trouvé")
        return {"success": True, "message": "Article supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting article: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la suppression")


# ===================== INCLUDE ROUTER AT END =====================
# This must be at the end to include all routes defined above
app.include_router(api_router)
